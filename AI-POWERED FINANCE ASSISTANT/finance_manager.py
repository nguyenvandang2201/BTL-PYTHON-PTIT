import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import calendar
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import hashlib

# Import ChatBot module
try:
    from chatbot import FinanceChatBot
    CHATBOT_AVAILABLE = True
except ImportError:
    CHATBOT_AVAILABLE = False
    print("Cảnh báo: Không thể import ChatBot. Vui lòng cài đặt: pip install google-generativeai")

# Import AI Auto Input module
try:
    from ai_auto_input import AIAutoInput
    AI_AUTO_INPUT_AVAILABLE = True
except ImportError:
    AI_AUTO_INPUT_AVAILABLE = False
    print("Cảnh báo: Không thể import AI Auto Input")

# Import Receipt OCR module
try:
    from receipt_ocr import ReceiptOCR
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    print("Cảnh báo: Không thể import Receipt OCR. Vui lòng cài đặt: pip install pillow")

# Import Gold Price module
try:
    from gold_price import GoldPriceAPI
    GOLD_PRICE_AVAILABLE = True
except ImportError:
    GOLD_PRICE_AVAILABLE = False
    print("Cảnh báo: Không thể import Gold Price API")

class FinanceManager:
    def __init__(self, root, user_id):
        self.root = root
        self.user_id = user_id
        self.root.title("Quản Lý Chi Tiêu Cá Nhân")
        self.root.geometry("1200x700")
        self.root.configure(bg="#f0f0f0")

        # Biến theo dõi trạng thái sắp xếp
        self.sort_ascending = True  # True: tăng dần, False: giảm dần

        # Khởi tạo database
        self.init_database()
        
        # Khởi tạo ChatBot
        if CHATBOT_AVAILABLE:
            try:
                self.chatbot = FinanceChatBot(user_id, self.conn)
            except Exception as e:
                print(f"Lỗi khởi tạo ChatBot: {e}")
                self.chatbot = None
        else:
            self.chatbot = None
        
        # Khởi tạo AI Auto Input
        if AI_AUTO_INPUT_AVAILABLE:
            try:
                self.ai_auto_input = AIAutoInput()
            except Exception as e:
                print(f"Lỗi khởi tạo AI Auto Input: {e}")
                self.ai_auto_input = None
        else:
            self.ai_auto_input = None
        
        # Khởi tạo Receipt OCR
        if OCR_AVAILABLE:
            try:
                self.receipt_ocr = ReceiptOCR()
            except Exception as e:
                print(f"Lỗi khởi tạo Receipt OCR: {e}")
                self.receipt_ocr = None
        else:
            self.receipt_ocr = None
        
        # Khởi tạo Gold Price API
        if GOLD_PRICE_AVAILABLE:
            try:
                self.gold_api = GoldPriceAPI()
            except Exception as e:
                print(f"Lỗi khởi tạo Gold Price API: {e}")
                self.gold_api = None
        else:
            self.gold_api = None

        # Tạo giao diện
        self.create_widgets()

        # Load dữ liệu ban đầu
        self.load_transactions()

    def init_database(self):
        """Khởi tạo cơ sở dữ liệu SQLite"""
        self.conn = sqlite3.connect('finance.db')
        self.cursor = self.conn.cursor()

        # Tạo bảng người dùng
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password TEXT NOT NULL
            )
        ''')

        # Tạo bảng giao dịch
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                type TEXT NOT NULL,
                category TEXT NOT NULL,
                amount REAL NOT NULL,
                description TEXT,
                date TEXT NOT NULL,
                user_id INTEGER,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')

        # Tạo bảng danh mục
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                type TEXT NOT NULL
            )
        ''')

        # Tạo bảng hạn mức chi tiêu
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS budget_limits (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                month INTEGER NOT NULL,
                year INTEGER NOT NULL,
                limit_amount REAL NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users (id),
                UNIQUE(user_id, month, year)
            )
        ''')

        # Thêm các danh mục mặc định nếu chưa có
        self.cursor.execute('SELECT COUNT(*) FROM categories')
        if self.cursor.fetchone()[0] == 0:
            default_categories = [
                ('Lương', 'income'),
                ('Thưởng', 'income'),
                ('Đầu tư', 'income'),
                ('Khác', 'income'),
                ('Ăn uống', 'expense'),
                ('Đi lại', 'expense'),
                ('Giải trí', 'expense'),
                ('Mua sắm', 'expense'),
                ('Hóa đơn', 'expense'),
                ('Y tế', 'expense'),
                ('Giáo dục', 'expense'),
                ('Khác', 'expense')
            ]
            self.cursor.executemany('INSERT INTO categories (name, type) VALUES (?, ?)',
                                   default_categories)

        self.conn.commit()

    def create_widgets(self):
        """Tạo giao diện người dùng"""
        # Frame chính
        main_frame = tk.Frame(self.root, bg="#f0f0f0")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Frame trái - Nhập liệu
        left_frame = tk.LabelFrame(main_frame, text="Nhập Giao Dịch",
                                   bg="white", font=("Arial", 12, "bold"),
                                   padx=15, pady=15)
        left_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        
        # Cấu hình grid để gold_frame có thể mở rộng
        left_frame.grid_rowconfigure(8, weight=1)
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_columnconfigure(1, weight=1)

        # Loại giao dịch
        tk.Label(left_frame, text="Loại:", bg="white", font=("Arial", 10)).grid(
            row=0, column=0, sticky="w", pady=5)
        self.type_var = tk.StringVar(value="expense")
        type_frame = tk.Frame(left_frame, bg="white")
        type_frame.grid(row=0, column=1, sticky="w", pady=5)
        tk.Radiobutton(type_frame, text="Thu nhập", variable=self.type_var,
                      value="income", bg="white", command=self.update_categories).pack(side=tk.LEFT)
        tk.Radiobutton(type_frame, text="Chi tiêu", variable=self.type_var,
                      value="expense", bg="white", command=self.update_categories).pack(side=tk.LEFT)

        # Danh mục
        tk.Label(left_frame, text="Danh mục:", bg="white", font=("Arial", 10)).grid(
            row=1, column=0, sticky="w", pady=5)
        self.category_var = tk.StringVar()
        self.category_combo = ttk.Combobox(left_frame, textvariable=self.category_var,
                                          state="readonly", width=25)
        self.category_combo.grid(row=1, column=1, sticky="w", pady=5)
        self.update_categories()

        # Số tiền
        tk.Label(left_frame, text="Số tiền:", bg="white", font=("Arial", 10)).grid(
            row=2, column=0, sticky="w", pady=5)
        self.amount_entry = tk.Entry(left_frame, width=27, font=("Arial", 10))
        self.amount_entry.grid(row=2, column=1, sticky="w", pady=5)

        # Mô tả
        tk.Label(left_frame, text="Mô tả:", bg="white", font=("Arial", 10)).grid(
            row=3, column=0, sticky="w", pady=5)
        self.description_entry = tk.Entry(left_frame, width=27, font=("Arial", 10))
        self.description_entry.grid(row=3, column=1, sticky="w", pady=5)

        # Ngày
        tk.Label(left_frame, text="Ngày:", bg="white", font=("Arial", 10)).grid(
            row=4, column=0, sticky="w", pady=5)
        date_frame = tk.Frame(left_frame, bg="white")
        date_frame.grid(row=4, column=1, sticky="w", pady=5)

        self.day_var = tk.StringVar(value=str(datetime.now().day))
        self.month_var = tk.StringVar(value=str(datetime.now().month))
        self.year_var = tk.StringVar(value=str(datetime.now().year))

        ttk.Combobox(date_frame, textvariable=self.day_var, width=3,
                    values=[str(i) for i in range(1, 32)], state="readonly").pack(side=tk.LEFT, padx=2)
        tk.Label(date_frame, text="/", bg="white").pack(side=tk.LEFT)
        ttk.Combobox(date_frame, textvariable=self.month_var, width=3,
                    values=[str(i) for i in range(1, 13)], state="readonly").pack(side=tk.LEFT, padx=2)
        tk.Label(date_frame, text="/", bg="white").pack(side=tk.LEFT)
        ttk.Combobox(date_frame, textvariable=self.year_var, width=5,
                    values=[str(i) for i in range(2020, 2031)], state="readonly").pack(side=tk.LEFT, padx=2)

        # Frame chứa các nút với kích thước đồng nhất
        # Nút Thêm Giao Dịch
        tk.Button(left_frame, text="➕ Thêm Giao Dịch", command=self.add_transaction,
                 bg="#4CAF50", fg="white", font=("Arial", 10, "bold"),
                 cursor="hand2", width=22, height=2).grid(row=5, column=0, columnspan=2, pady=5, sticky="ew")
        
        # Nút Nhập bằng AI
        tk.Button(left_frame, text="🤖 Nhập bằng AI", command=self.open_ai_auto_input,
                 bg="#FF5722", fg="white", font=("Arial", 10, "bold"),
                 cursor="hand2", width=22, height=2).grid(row=6, column=0, columnspan=2, pady=5, sticky="ew")
        
        # Nút Quét Hóa Đơn
        tk.Button(left_frame, text="📷 Quét Hóa Đơn", command=self.open_receipt_ocr,
                 bg="#FF9800", fg="white", font=("Arial", 10, "bold"),
                 cursor="hand2", width=22, height=2).grid(row=7, column=0, columnspan=2, pady=5, sticky="ew")

        # Frame chứa Giá vàng và Bitcoin (xếp dọc)
        price_container = tk.Frame(left_frame, bg="white")
        price_container.grid(row=8, column=0, columnspan=2, pady=10, sticky="nsew")
        
        # Frame Giá Vàng
        gold_frame = tk.LabelFrame(price_container, text="💰 Giá Vàng Hôm Nay",
                                   bg="white", font=("Arial", 9, "bold"),
                                   fg="#FF9800", padx=8, pady=5)
        gold_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        
        # Nội dung giá vàng
        self.gold_price_label = tk.Label(gold_frame, text="⏳ Đang tải...", 
                                         bg="white", font=("Arial", 10, "bold"),
                                         justify=tk.CENTER, fg="#333", height=4)
        self.gold_price_label.pack(fill=tk.BOTH, expand=True)
        
        # Nút refresh giá vàng
        refresh_gold_btn = tk.Button(gold_frame, text="🔄 Cập nhật",
                                     command=self.update_gold_price,
                                     bg="#FFC107", fg="white",
                                     font=("Arial", 7, "bold"),
                                     cursor="hand2", width=10)
        refresh_gold_btn.pack(pady=(3, 0))
        
        # Frame Giá Bitcoin
        btc_frame = tk.LabelFrame(price_container, text="₿ Giá Bitcoin Hiện Tại",
                                  bg="white", font=("Arial", 9, "bold"),
                                  fg="#F7931A", padx=8, pady=5)
        btc_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
        
        # Nội dung giá Bitcoin
        self.btc_price_label = tk.Label(btc_frame, text="⏳ Đang tải...", 
                                        bg="white", font=("Arial", 10, "bold"),
                                        justify=tk.CENTER, fg="#333", height=4)
        self.btc_price_label.pack(fill=tk.BOTH, expand=True)
        
        # Nút refresh giá Bitcoin
        refresh_btc_btn = tk.Button(btc_frame, text="🔄 Cập nhật",
                                    command=self.update_btc_price,
                                    bg="#F7931A", fg="white",
                                    font=("Arial", 7, "bold"),
                                    cursor="hand2", width=10)
        refresh_btc_btn.pack(pady=(3, 0))
        
        # Tải giá lần đầu
        self.update_gold_price()
        self.update_btc_price()
        
        # Auto-refresh mỗi 5 phút
        self.schedule_gold_price_update()
        self.schedule_btc_price_update()

        # Frame giữa - Danh sách giao dịch
        middle_frame = tk.LabelFrame(main_frame, text="Danh Sách Giao Dịch",
                                     bg="white", font=("Arial", 12, "bold"),
                                     padx=10, pady=10)
        middle_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)

        # Bộ lọc
        filter_frame = tk.Frame(middle_frame, bg="white")
        filter_frame.pack(fill=tk.X, pady=5)

        # Hàng 1: Loại và Danh mục
        filter_row1 = tk.Frame(filter_frame, bg="white")
        filter_row1.pack(fill=tk.X, pady=2)

        tk.Label(filter_row1, text="Loại:", bg="white").pack(side=tk.LEFT, padx=5)
        self.filter_type_var = tk.StringVar(value="Tất cả")
        self.filter_type_combo = ttk.Combobox(filter_row1, textvariable=self.filter_type_var, width=10,
                    values=["Tất cả", "Thu nhập", "Chi tiêu"],
                    state="readonly")
        self.filter_type_combo.pack(side=tk.LEFT, padx=5)
        self.filter_type_combo.bind("<<ComboboxSelected>>", self.update_filter_categories)

        tk.Label(filter_row1, text="Danh mục:", bg="white").pack(side=tk.LEFT, padx=5)
        self.filter_category_var = tk.StringVar(value="Tất cả")
        self.filter_category_combo = ttk.Combobox(filter_row1, textvariable=self.filter_category_var, width=12,
                    state="readonly")
        self.filter_category_combo.pack(side=tk.LEFT, padx=5)
        self.update_filter_categories()

        # Hàng 2: Tháng, Năm và nút Lọc
        filter_row2 = tk.Frame(filter_frame, bg="white")
        filter_row2.pack(fill=tk.X, pady=2)

        tk.Label(filter_row2, text="Tháng:", bg="white").pack(side=tk.LEFT, padx=5)
        self.filter_month_var = tk.StringVar(value=str(datetime.now().month))
        ttk.Combobox(filter_row2, textvariable=self.filter_month_var, width=5,
                    values=["Tất cả"] + [str(i) for i in range(1, 13)],
                    state="readonly").pack(side=tk.LEFT, padx=5)

        tk.Label(filter_row2, text="Năm:", bg="white").pack(side=tk.LEFT, padx=5)
        self.filter_year_var = tk.StringVar(value=str(datetime.now().year))
        ttk.Combobox(filter_row2, textvariable=self.filter_year_var, width=8,
                    values=["Tất cả"] + [str(i) for i in range(2020, 2031)],
                    state="readonly").pack(side=tk.LEFT, padx=5)

        # Hàng 2.5: Lọc theo khoảng ngày
        filter_row2_5 = tk.Frame(filter_frame, bg="white")
        filter_row2_5.pack(fill=tk.X, pady=2)

        tk.Label(filter_row2_5, text="Từ ngày:", bg="white").pack(side=tk.LEFT, padx=5)
        self.filter_date_from_var = tk.StringVar()
        from_date_entry = tk.Entry(filter_row2_5, textvariable=self.filter_date_from_var, 
                                   width=12, font=("Arial", 9))
        from_date_entry.pack(side=tk.LEFT, padx=5)
        tk.Label(filter_row2_5, text="(dd/mm/yyyy)", bg="white", 
                font=("Arial", 8), fg="#666").pack(side=tk.LEFT, padx=2)

        tk.Label(filter_row2_5, text="Đến ngày:", bg="white").pack(side=tk.LEFT, padx=5)
        self.filter_date_to_var = tk.StringVar()
        to_date_entry = tk.Entry(filter_row2_5, textvariable=self.filter_date_to_var, 
                                width=12, font=("Arial", 9))
        to_date_entry.pack(side=tk.LEFT, padx=5)
        tk.Label(filter_row2_5, text="(dd/mm/yyyy)", bg="white", 
                font=("Arial", 8), fg="#666").pack(side=tk.LEFT, padx=2)

        # Hàng 3: Tìm kiếm theo mô tả
        filter_row3 = tk.Frame(filter_frame, bg="white")
        filter_row3.pack(fill=tk.X, pady=2)

        tk.Label(filter_row3, text="Tìm kiếm:", bg="white").pack(side=tk.LEFT, padx=5)
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(filter_row3, textvariable=self.search_var, width=30, font=("Arial", 9))
        search_entry.pack(side=tk.LEFT, padx=5)
        search_entry.bind('<Return>', lambda e: self.load_transactions())  # Tìm kiếm khi nhấn Enter

        tk.Button(filter_row3, text="Lọc", command=self.load_transactions,
                 bg="#2196F3", fg="white", cursor="hand2", padx=15).pack(side=tk.LEFT, padx=5)

        tk.Button(filter_row3, text="Xóa Bộ Lọc", command=self.reset_filters,
                 bg="#9E9E9E", fg="white", cursor="hand2", padx=10).pack(side=tk.LEFT, padx=5)

        tk.Button(filter_row3, text="Xóa Giao Dịch", command=self.delete_transaction,
                 bg="#f44336", fg="white", cursor="hand2", padx=10).pack(side=tk.LEFT, padx=5)

        # Bảng giao dịch
        tree_frame = tk.Frame(middle_frame, bg="white")
        tree_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("ID", "Loại", "Danh mục", "Số tiền", "Mô tả", "Ngày")
        self.transaction_tree = ttk.Treeview(tree_frame, columns=columns,
                                            show="headings", height=15)

        # Định nghĩa tiêu đề cột
        self.transaction_tree.heading("ID", text="ID")
        self.transaction_tree.heading("Loại", text="Loại")
        self.transaction_tree.heading("Danh mục", text="Danh mục")
        self.transaction_tree.heading("Số tiền", text="Số tiền (VNĐ) ↕", 
                                     command=self.sort_by_amount)
        self.transaction_tree.heading("Mô tả", text="Mô tả")
        self.transaction_tree.heading("Ngày", text="Ngày")

        # Định dạng cột
        self.transaction_tree.column("ID", width=40, anchor="center")
        self.transaction_tree.column("Loại", width=80, anchor="center")
        self.transaction_tree.column("Danh mục", width=100, anchor="center")
        self.transaction_tree.column("Số tiền", width=120, anchor="e")
        self.transaction_tree.column("Mô tả", width=150)
        self.transaction_tree.column("Ngày", width=90, anchor="center")

        # Thanh cuộn
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL,
                                 command=self.transaction_tree.yview)
        self.transaction_tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.transaction_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Frame phải - Thống kê và biểu đồ
        right_frame = tk.Frame(main_frame, bg="white")
        right_frame.grid(row=0, column=2, sticky="nsew", padx=5, pady=5)

        # Nút ChatBot AI - Đặt lên đầu tiên
        ai_frame = tk.LabelFrame(right_frame, text="🤖 Trợ Lý AI",
                                bg="white", font=("Arial", 12, "bold"),
                                padx=10, pady=10)
        ai_frame.pack(fill=tk.X, pady=5)
        
        tk.Button(ai_frame, text="💬 Trợ Lý Tài Chính AI",
                 command=self.open_chatbot,
                 bg="#4285F4", fg="white", font=("Arial", 11, "bold"),
                 cursor="hand2", width=22, height=2).pack(pady=3)

        # Thống kê
        stats_frame = tk.LabelFrame(right_frame, text="Thống Kê",
                                   bg="white", font=("Arial", 12, "bold"),
                                   padx=10, pady=10)
        stats_frame.pack(fill=tk.X, pady=5)

        self.income_label = tk.Label(stats_frame, text="Tổng thu nhập: 0 VNĐ",
                                     bg="white", fg="green", font=("Arial", 11, "bold"))
        self.income_label.pack(anchor="w", pady=3)

        self.expense_label = tk.Label(stats_frame, text="Tổng chi tiêu: 0 VNĐ",
                                      bg="white", fg="red", font=("Arial", 11, "bold"))
        self.expense_label.pack(anchor="w", pady=3)

        self.balance_label = tk.Label(stats_frame, text="Số dư: 0 VNĐ",
                                      bg="white", fg="blue", font=("Arial", 11, "bold"))
        self.balance_label.pack(anchor="w", pady=3)

        # Quản lý giao dịch
        transaction_mgmt_frame = tk.LabelFrame(right_frame, text="📋 Quản Lý Giao Dịch",
                                             bg="white", font=("Arial", 11, "bold"),
                                             padx=8, pady=8)
        transaction_mgmt_frame.pack(fill=tk.X, pady=5)

        tk.Button(transaction_mgmt_frame, text="📥 Nhập từ Excel",
                 command=self.import_from_excel,
                 bg="#008000", fg="white", font=("Arial", 9),
                 cursor="hand2", width=20).pack(pady=2)

        tk.Button(transaction_mgmt_frame, text="📤 Xuất Excel",
                 command=self.export_to_excel,
                 bg="#00796B", fg="white", font=("Arial", 9),
                 cursor="hand2", width=20).pack(pady=2)

        tk.Button(transaction_mgmt_frame, text="📄 Xuất PDF",
                 command=self.export_to_pdf,
                 bg="#673AB7", fg="white", font=("Arial", 9),
                 cursor="hand2", width=20).pack(pady=2)

        # Nút biểu đồ - Thu gọn
        chart_frame = tk.LabelFrame(right_frame, text="📊 Biểu Đồ",
                                   bg="white", font=("Arial", 11, "bold"),
                                   padx=8, pady=8)
        chart_frame.pack(fill=tk.X, pady=5)

        tk.Button(chart_frame, text="📈 Theo Danh Mục",
                 command=self.show_category_chart,
                 bg="#FF9800", fg="white", font=("Arial", 9),
                 cursor="hand2", width=20).pack(pady=2)

        tk.Button(chart_frame, text="📊 Theo Tháng",
                 command=self.show_monthly_chart,
                 bg="#9C27B0", fg="white", font=("Arial", 9),
                 cursor="hand2", width=20).pack(pady=2)

        tk.Button(chart_frame, text="📉 Theo Năm",
                 command=self.show_yearly_chart,
                 bg="#3F51B5", fg="white", font=("Arial", 9),
                 cursor="hand2", width=20).pack(pady=2)

        # Quản lý danh mục
        category_mgmt_frame = tk.LabelFrame(right_frame, text="⚙️ Quản Lý",
                                   bg="white", font=("Arial", 11, "bold"),
                                   padx=8, pady=8)
        category_mgmt_frame.pack(fill=tk.X, pady=5)

        tk.Button(category_mgmt_frame, text="📝 Quản Lý Danh Mục",
                 command=self.manage_categories,
                 bg="#607D8B", fg="white", font=("Arial", 9),
                 cursor="hand2", width=20).pack(pady=2)

        tk.Button(category_mgmt_frame, text="💰 Đặt Hạn Mức Chi Tiêu",
                 command=self.set_budget_limit,
                 bg="#E91E63", fg="white", font=("Arial", 9),
                 cursor="hand2", width=20).pack(pady=2)

        # Bảng thông báo hạn mức
        budget_info_frame = tk.LabelFrame(right_frame, text="💰 Hạn Mức Tháng Này",
                                         bg="white", font=("Arial", 11, "bold"),
                                         padx=8, pady=8)
        budget_info_frame.pack(fill=tk.X, pady=5)

        # Label hiển thị thông tin hạn mức
        self.budget_info_label = tk.Label(budget_info_frame, text="", 
                                         bg="white", font=("Arial", 9),
                                         justify=tk.LEFT, anchor="w")
        self.budget_info_label.pack(fill=tk.X, pady=5)
        
        # Load thông tin hạn mức ban đầu
        self.update_budget_info_display()

        # Cấu hình grid
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=2)
        main_frame.grid_columnconfigure(2, weight=1)
        main_frame.grid_rowconfigure(0, weight=1)


    def update_categories(self):
        """Cập nhật danh sách danh mục theo loại"""
        trans_type = self.type_var.get()
        self.cursor.execute('SELECT name FROM categories WHERE type = ?', (trans_type,))
        categories = [row[0] for row in self.cursor.fetchall()]
        self.category_combo['values'] = categories
        if categories:
            self.category_combo.current(0)

    def update_filter_categories(self, event=None):
        """Cập nhật danh sách danh mục trong bộ lọc theo loại được chọn"""
        filter_type = self.filter_type_var.get()

        if filter_type == "Tất cả":
            # Lấy tất cả danh mục
            self.cursor.execute('SELECT DISTINCT name FROM categories ORDER BY name')
            categories = ["Tất cả"] + [row[0] for row in self.cursor.fetchall()]
        else:
            # Chuyển đổi từ tiếng Việt sang tiếng Anh
            trans_type = "income" if filter_type == "Thu nhập" else "expense"
            self.cursor.execute('SELECT name FROM categories WHERE type = ? ORDER BY name', (trans_type,))
            categories = ["Tất cả"] + [row[0] for row in self.cursor.fetchall()]

        self.filter_category_combo['values'] = categories
        self.filter_category_var.set("Tất cả")

    def reset_filters(self):
        """Xóa tất cả bộ lọc và hiển thị toàn bộ dữ liệu"""
        self.filter_type_var.set("Tất cả")
        self.filter_category_var.set("Tất cả")
        self.filter_month_var.set("Tất cả")
        self.filter_year_var.set("Tất cả")
        self.filter_date_from_var.set("")  # Xóa ngày bắt đầu
        self.filter_date_to_var.set("")    # Xóa ngày kết thúc
        self.search_var.set("")  # Xóa tìm kiếm
        self.update_filter_categories()
        self.load_transactions()

    def add_transaction(self):
        """Thêm giao dịch mới"""
        try:
            trans_type = self.type_var.get()
            category = self.category_var.get()
            amount = float(self.amount_entry.get())
            description = self.description_entry.get()
            date = f"{self.year_var.get()}-{self.month_var.get().zfill(2)}-{self.day_var.get().zfill(2)}"

            if not category:
                messagebox.showerror("Lỗi", "Vui lòng chọn danh mục!")
                return

            if amount <= 0:
                messagebox.showerror("Lỗi", "Số tiền phải lớn hơn 0!")
                return

            self.cursor.execute('''
                INSERT INTO transactions (type, category, amount, description, date, user_id)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (trans_type, category, amount, description, date, self.user_id))

            self.conn.commit()
            messagebox.showinfo("Thành công", "Đã thêm giao dịch!")

            # Reset form
            self.amount_entry.delete(0, tk.END)
            self.description_entry.delete(0, tk.END)

            # Cập nhật danh sách
            self.load_transactions()

        except ValueError:
            messagebox.showerror("Lỗi", "Số tiền không hợp lệ!")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra: {str(e)}")

    def load_transactions(self):
        """Tải danh sách giao dịch"""
        # Xóa dữ liệu cũ
        for item in self.transaction_tree.get_children():
            self.transaction_tree.delete(item)

        # Lấy bộ lọc
        filter_type = self.filter_type_var.get()
        filter_category = self.filter_category_var.get()
        filter_month = self.filter_month_var.get()
        filter_year = self.filter_year_var.get()
        search_keyword = self.search_var.get().strip()

        # Truy vấn - chỉ lấy giao dịch của user hiện tại
        query = 'SELECT * FROM transactions WHERE user_id = ?'
        params = [self.user_id]

        # Lọc theo loại
        if filter_type != "Tất cả":
            trans_type = "income" if filter_type == "Thu nhập" else "expense"
            query += ' AND type = ?'
            params.append(trans_type)

        # Lọc theo danh mục
        if filter_category != "Tất cả":
            query += ' AND category = ?'
            params.append(filter_category)

        # Lọc theo tháng
        if filter_month != "Tất cả":
            query += ' AND strftime("%m", date) = ?'
            params.append(filter_month.zfill(2))

        # Lọc theo năm
        if filter_year != "Tất cả":
            query += ' AND strftime("%Y", date) = ?'
            params.append(filter_year)

        # Lọc theo khoảng ngày
        date_from = self.filter_date_from_var.get().strip()
        date_to = self.filter_date_to_var.get().strip()
        
        if date_from:
            try:
                # Chuyển đổi từ dd/mm/yyyy sang yyyy-mm-dd
                date_parts = date_from.split('/')
                if len(date_parts) == 3:
                    formatted_date = f"{date_parts[2]}-{date_parts[1].zfill(2)}-{date_parts[0].zfill(2)}"
                    query += ' AND date >= ?'
                    params.append(formatted_date)
            except:
                pass  # Bỏ qua nếu định dạng không đúng

        if date_to:
            try:
                # Chuyển đổi từ dd/mm/yyyy sang yyyy-mm-dd
                date_parts = date_to.split('/')
                if len(date_parts) == 3:
                    formatted_date = f"{date_parts[2]}-{date_parts[1].zfill(2)}-{date_parts[0].zfill(2)}"
                    query += ' AND date <= ?'
                    params.append(formatted_date)
            except:
                pass  # Bỏ qua nếu định dạng không đúng

        # Tìm kiếm theo mô tả
        if search_keyword:
            query += ' AND (description LIKE ? OR category LIKE ?)'
            params.append(f'%{search_keyword}%')
            params.append(f'%{search_keyword}%')

        query += ' ORDER BY date DESC'

        self.cursor.execute(query, params)
        transactions = self.cursor.fetchall()

        # Thêm vào bảng
        total_income = 0
        total_expense = 0

        for trans in transactions:
            # Unpack đúng số cột (7 cột: id, type, category, amount, description, date, user_id)
            trans_id, trans_type, category, amount, description, date, user_id = trans
            type_text = "Thu nhập" if trans_type == "income" else "Chi tiêu"
            amount_text = f"{amount:,.0f}"

            # Đổi định dạng ngày
            date_parts = date.split('-')
            date_formatted = f"{date_parts[2]}/{date_parts[1]}/{date_parts[0]}"

            self.transaction_tree.insert("", 0, values=(
                trans_id, type_text, category, amount_text, description, date_formatted
            ))

            if trans_type == "income":
                total_income += amount
            else:
                total_expense += amount

        # Cập nhật thống kê
        balance = total_income - total_expense
        self.income_label.config(text=f"Tổng thu nhập: {total_income:,.0f} VNĐ")
        self.expense_label.config(text=f"Tổng chi tiêu: {total_expense:,.0f} VNĐ")
        self.balance_label.config(text=f"Số dư: {balance:,.0f} VNĐ")
        
        # Cập nhật hiển thị hạn mức
        self.update_budget_info_display()
        
        # Kiểm tra cảnh báo hạn mức
        self.check_budget_warning()

    def check_budget_warning(self):
        """Kiểm tra và hiển thị cảnh báo nếu vượt hạn mức chi tiêu"""
        # Lấy tháng và năm hiện tại
        now = datetime.now()
        current_month = now.month
        current_year = now.year
        
        # Lấy hạn mức của tháng hiện tại
        self.cursor.execute('''
            SELECT limit_amount FROM budget_limits 
            WHERE user_id = ? AND month = ? AND year = ?
        ''', (self.user_id, current_month, current_year))
        
        result = self.cursor.fetchone()
        
        if not result:
            return  # Không có hạn mức được đặt
        
        limit_amount = result[0]
        
        # Tính tổng chi tiêu trong tháng hiện tại
        month_str = str(current_month).zfill(2)
        year_str = str(current_year)
        
        self.cursor.execute('''
            SELECT COALESCE(SUM(amount), 0) FROM transactions 
            WHERE user_id = ? AND type = "expense" 
            AND strftime("%m", date) = ? AND strftime("%Y", date) = ?
        ''', (self.user_id, month_str, year_str))
        
        total_expense = self.cursor.fetchone()[0]
        
        # Tính phần trăm đã chi tiêu
        percentage = (total_expense / limit_amount * 100) if limit_amount > 0 else 0
        
        # Hiển thị cảnh báo
        if total_expense > limit_amount:
            over_amount = total_expense - limit_amount
            messagebox.showwarning(
                "⚠️ Cảnh Báo Hạn Mức",
                f"Bạn đã VƯỢT hạn mức chi tiêu tháng {current_month}/{current_year}!\n\n"
                f"Hạn mức: {limit_amount:,.0f} VNĐ\n"
                f"Đã chi tiêu: {total_expense:,.0f} VNĐ ({percentage:.1f}%)\n"
                f"Vượt: {over_amount:,.0f} VNĐ"
            )
        elif percentage >= 90:
            remaining = limit_amount - total_expense
            messagebox.showwarning(
                "⚠️ Cảnh Báo Hạn Mức",
                f"Chi tiêu của bạn đã đạt {percentage:.1f}% hạn mức tháng {current_month}/{current_year}!\n\n"
                f"Hạn mức: {limit_amount:,.0f} VNĐ\n"
                f"Đã chi tiêu: {total_expense:,.0f} VNĐ\n"
                f"Còn lại: {remaining:,.0f} VNĐ"
            )
        elif percentage >= 80:
            remaining = limit_amount - total_expense
            messagebox.showinfo(
                "ℹ️ Thông Báo Hạn Mức",
                f"Chi tiêu của bạn đã đạt {percentage:.1f}% hạn mức tháng {current_month}/{current_year}\n\n"
                f"Hạn mức: {limit_amount:,.0f} VNĐ\n"
                f"Đã chi tiêu: {total_expense:,.0f} VNĐ\n"
                f"Còn lại: {remaining:,.0f} VNĐ"
            )

    def update_budget_info_display(self):
        """Cập nhật hiển thị thông tin hạn mức trong bảng thông báo"""
        # Lấy tháng và năm hiện tại
        now = datetime.now()
        current_month = now.month
        current_year = now.year
        
        # Lấy hạn mức của tháng hiện tại
        self.cursor.execute('''
            SELECT limit_amount FROM budget_limits 
            WHERE user_id = ? AND month = ? AND year = ?
        ''', (self.user_id, current_month, current_year))
        
        result = self.cursor.fetchone()
        
        if not result:
            # Không có hạn mức
            self.budget_info_label.config(
                text="Chưa đặt hạn mức\ncho tháng này",
                fg="#999"
            )
            return
        
        limit_amount = result[0]
        
        # Tính tổng chi tiêu trong tháng hiện tại
        month_str = str(current_month).zfill(2)
        year_str = str(current_year)
        
        self.cursor.execute('''
            SELECT COALESCE(SUM(amount), 0) FROM transactions 
            WHERE user_id = ? AND type = "expense" 
            AND strftime("%m", date) = ? AND strftime("%Y", date) = ?
        ''', (self.user_id, month_str, year_str))
        
        total_expense = self.cursor.fetchone()[0]
        
        # Tính phần trăm và còn lại
        percentage = (total_expense / limit_amount * 100) if limit_amount > 0 else 0
        remaining = limit_amount - total_expense
        
        # Tạo text hiển thị
        info_text = f"Hạn mức: {limit_amount:,.0f} VNĐ\n"
        info_text += f"Đã chi: {total_expense:,.0f} VNĐ ({percentage:.1f}%)\n"
        info_text += f"Còn lại: {remaining:,.0f} VNĐ"
        
        # Đổi màu theo tình trạng
        if total_expense > limit_amount:
            color = "#f44336"  # Đỏ - vượt hạn mức
        elif percentage >= 90:
            color = "#FF5722"  # Cam đậm - gần hạn mức
        elif percentage >= 80:
            color = "#FF9800"  # Cam - cảnh báo
        elif percentage >= 50:
            color = "#FFC107"  # Vàng - bình thường
        else:
            color = "#4CAF50"  # Xanh - tốt
        
        self.budget_info_label.config(text=info_text, fg=color)

    def sort_by_amount(self):
        """Sắp xếp danh sách giao dịch theo số tiền"""
        # Lấy tất cả các mục hiện tại
        items = []
        for item in self.transaction_tree.get_children():
            values = self.transaction_tree.item(item)['values']
            # Chuyển số tiền từ chuỗi có dấu phân cách về số
            amount_str = values[3].replace(',', '')
            amount = float(amount_str)
            items.append((item, values, amount))
        
        # Sắp xếp theo số tiền
        items.sort(key=lambda x: x[2], reverse=not self.sort_ascending)
        
        # Xóa tất cả các mục
        for item in self.transaction_tree.get_children():
            self.transaction_tree.delete(item)
        
        # Thêm lại theo thứ tự đã sắp xếp
        for item, values, amount in items:
            self.transaction_tree.insert("", "end", values=values)
        
        # Đổi trạng thái sắp xếp và cập nhật biểu tượng
        self.sort_ascending = not self.sort_ascending
        sort_symbol = "↑" if self.sort_ascending else "↓"
        self.transaction_tree.heading("Số tiền", text=f"Số tiền (VNĐ) {sort_symbol}")

    def delete_transaction(self):
        """Xóa giao dịch được chọn"""
        selected = self.transaction_tree.selection()
        if not selected:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn giao dịch để xóa!")
            return

        if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa giao dịch này?"):
            for item in selected:
                trans_id = self.transaction_tree.item(item)['values'][0]
                self.cursor.execute('DELETE FROM transactions WHERE id = ?', (trans_id,))

            self.conn.commit()
            messagebox.showinfo("Thành công", "Đã xóa giao dịch!")
            self.load_transactions()

    def show_category_chart(self):
        """Hiển thị biểu đồ theo danh mục"""
        filter_month = self.filter_month_var.get()
        filter_year = self.filter_year_var.get()

        query = 'SELECT category, SUM(amount) FROM transactions WHERE type = "expense"'
        params = []

        if filter_month != "Tất cả":
            query += ' AND strftime("%m", date) = ?'
            params.append(filter_month.zfill(2))

        if filter_year != "Tất cả":
            query += ' AND strftime("%Y", date) = ?'
            params.append(filter_year)

        query += ' GROUP BY category'

        self.cursor.execute(query, params)
        data = self.cursor.fetchall()

        if not data:
            messagebox.showinfo("Thông báo", "Không có dữ liệu để hiển thị!")
            return

        categories = [row[0] for row in data]
        amounts = [row[1] for row in data]

        # Tạo cửa sổ mới
        chart_window = tk.Toplevel(self.root)
        chart_window.title("Biểu Đồ Chi Tiêu Theo Danh Mục")
        chart_window.geometry("800x600")

        fig = Figure(figsize=(8, 6))
        ax = fig.add_subplot(111)

        colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40']
        ax.pie(amounts, labels=categories, autopct='%1.1f%%', startangle=90, colors=colors)
        ax.set_title('Chi Tiêu Theo Danh Mục', fontsize=14, fontweight='bold')

        canvas = FigureCanvasTkAgg(fig, master=chart_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def show_monthly_chart(self):
        """Hiển thị biểu đồ theo tháng"""
        filter_year = self.filter_year_var.get()

        if filter_year == "Tất cả":
            filter_year = str(datetime.now().year)

        # Lấy dữ liệu thu nhập và chi tiêu theo tháng
        months = list(range(1, 13))
        income_data = []
        expense_data = []

        for month in months:
            # Thu nhập
            self.cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) FROM transactions 
                WHERE type = "income" AND strftime("%m", date) = ? AND strftime("%Y", date) = ?
            ''', (str(month).zfill(2), filter_year))
            income_data.append(self.cursor.fetchone()[0])

            # Chi tiêu
            self.cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) FROM transactions 
                WHERE type = "expense" AND strftime("%m", date) = ? AND strftime("%Y", date) = ?
            ''', (str(month).zfill(2), filter_year))
            expense_data.append(self.cursor.fetchone()[0])

        # Tạo cửa sổ mới
        chart_window = tk.Toplevel(self.root)
        chart_window.title(f"Biểu Đồ Tài Chính Năm {filter_year}")
        chart_window.geometry("900x600")

        fig = Figure(figsize=(9, 6))
        ax = fig.add_subplot(111)

        x = range(1, 13)
        width = 0.35

        ax.bar([i - width/2 for i in x], income_data, width, label='Thu nhập', color='#4CAF50')
        ax.bar([i + width/2 for i in x], expense_data, width, label='Chi tiêu', color='#f44336')

        ax.set_xlabel('Tháng', fontsize=12)
        ax.set_ylabel('Số tiền (VNĐ)', fontsize=12)
        ax.set_title(f'Biểu Đồ Tài Chính Theo Tháng - Năm {filter_year}',
                     fontsize=14, fontweight='bold')
        ax.set_xticks(x)
        ax.legend()
        ax.grid(axis='y', alpha=0.3)

        canvas = FigureCanvasTkAgg(fig, master=chart_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def show_yearly_chart(self):
        """Hiển thị biểu đồ theo năm"""
        # Lấy danh sách các năm có dữ liệu
        self.cursor.execute('''
            SELECT DISTINCT strftime("%Y", date) FROM transactions 
            ORDER BY date
        ''')
        years = [row[0] for row in self.cursor.fetchall()]

        if not years:
            messagebox.showinfo("Thông báo", "Không có dữ liệu để hiển thị!")
            return

        income_data = []
        expense_data = []

        for year in years:
            # Thu nhập
            self.cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) FROM transactions 
                WHERE type = "income" AND strftime("%Y", date) = ?
            ''', (year,))
            income_data.append(self.cursor.fetchone()[0])

            # Chi tiêu
            self.cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) FROM transactions 
                WHERE type = "expense" AND strftime("%Y", date) = ?
            ''', (year,))
            expense_data.append(self.cursor.fetchone()[0])

        # Tạo cửa sổ mới
        chart_window = tk.Toplevel(self.root)
        chart_window.title("Biểu Đồ Tài Chính Theo Năm")
        chart_window.geometry("900x600")

        fig = Figure(figsize=(9, 6))
        ax = fig.add_subplot(111)

        x = range(len(years))
        width = 0.35

        ax.bar([i - width/2 for i in x], income_data, width, label='Thu nhập', color='#4CAF50')
        ax.bar([i + width/2 for i in x], expense_data, width, label='Chi tiêu', color='#f44336')

        ax.set_xlabel('Năm', fontsize=12)
        ax.set_ylabel('Số tiền (VNĐ)', fontsize=12)
        ax.set_title('Biểu Đồ Tài Chính Theo Năm', fontsize=14, fontweight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels(years)
        ax.legend()
        ax.grid(axis='y', alpha=0.3)

        canvas = FigureCanvasTkAgg(fig, master=chart_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def export_to_pdf(self):
        """Xuất danh sách giao dịch ra file PDF"""
        # Lấy bộ lọc hiện tại
        filter_month = self.filter_month_var.get()
        filter_year = self.filter_year_var.get()

        # Truy vấn dữ liệu
        query = 'SELECT * FROM transactions WHERE user_id = ?'
        params = [self.user_id]

        if filter_month != "Tất cả":
            query += ' AND strftime("%m", date) = ?'
            params.append(filter_month.zfill(2))

        if filter_year != "Tất cả":
            query += ' AND strftime("%Y", date) = ?'
            params.append(filter_year)

        query += ' ORDER BY date DESC'

        self.cursor.execute(query, params)
        transactions = self.cursor.fetchall()

        if not transactions:
            messagebox.showinfo("Thông báo", "Không có dữ liệu để xuất!")
            return

        # Chọn nơi lưu file
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=f"giao_dich_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )

        if not filename:
            return

        try:
            # Đăng ký font tiếng Việt
            font_name = 'Helvetica'
            font_bold = 'Helvetica-Bold'

            # Thử tải font DejaVu Sans từ thư mục hiện tại
            try:
                import os
                current_dir = os.path.dirname(os.path.abspath(__file__))
                dejavu_path = os.path.join(current_dir, 'DejaVuSans.ttf')
                dejavu_bold_path = os.path.join(current_dir, 'DejaVuSans-Bold.ttf')

                if os.path.exists(dejavu_path):
                    pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_path))
                    font_name = 'DejaVuSans'
                    if os.path.exists(dejavu_bold_path):
                        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_bold_path))
                        font_bold = 'DejaVuSans-Bold'
                    else:
                        font_bold = 'DejaVuSans'
                else:
                    # Thử tìm font Arial Unicode MS trong Windows
                    windows_font_path = r'C:\Windows\Fonts\arial.ttf'
                    if os.path.exists(windows_font_path):
                        pdfmetrics.registerFont(TTFont('Arial', windows_font_path))
                        font_name = 'Arial'
                        font_bold = 'Arial'
            except Exception as e:
                # Nếu không tìm thấy font, dùng Helvetica mặc định
                print(f"Không thể tải font tiếng Việt: {e}")

            # Tạo PDF
            doc = SimpleDocTemplate(filename, pagesize=A4)
            elements = []

            # Style
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1a237e'),
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName=font_bold
            )

            # Tiêu đề
            title_text = "DANH SÁCH GIAO DỊCH"
            if filter_month != "Tất cả" and filter_year != "Tất cả":
                title_text += f"<br/>Tháng {filter_month}/{filter_year}"
            elif filter_year != "Tất cả":
                title_text += f"<br/>Năm {filter_year}"

            title = Paragraph(title_text, title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.3*inch))

            # Tạo bảng dữ liệu
            data = [['STT', 'Loại', 'Danh mục', 'Số tiền (VNĐ)', 'Mô tả', 'Ngày']]

            total_income = 0
            total_expense = 0

            for idx, trans in enumerate(transactions, 1):
                trans_id, trans_type, category, amount, description, date, user_id = trans
                type_text = "Thu nhập" if trans_type == "income" else "Chi tiêu"
                amount_text = f"{amount:,.0f}"

                # Đổi định dạng ngày
                date_parts = date.split('-')
                date_formatted = f"{date_parts[2]}/{date_parts[1]}/{date_parts[0]}"

                data.append([
                    str(idx),
                    type_text,
                    category,
                    amount_text,
                    description or "",
                    date_formatted
                ])

                if trans_type == "income":
                    total_income += amount
                else:
                    total_expense += amount

            # Tạo bảng
            table = Table(data, colWidths=[0.6*inch, 1*inch, 1.2*inch, 1.3*inch, 2*inch, 1*inch])

            # Style cho bảng với font tiếng Việt
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_bold),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

                # Dữ liệu
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # STT
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Loại
                ('ALIGN', (2, 1), (2, -1), 'LEFT'),    # Danh mục
                ('ALIGN', (3, 1), (3, -1), 'RIGHT'),   # Số tiền
                ('ALIGN', (4, 1), (4, -1), 'LEFT'),    # Mô tả
                ('ALIGN', (5, 1), (5, -1), 'CENTER'),  # Ngày
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),

                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            elements.append(table)
            elements.append(Spacer(1, 0.3*inch))

            # Thống kê
            balance = total_income - total_expense
            summary_style = ParagraphStyle(
                'Summary',
                parent=styles['Normal'],
                fontSize=11,
                textColor=colors.HexColor('#1a237e'),
                spaceAfter=8,
                alignment=TA_RIGHT,
                fontName=font_bold
            )

            elements.append(Paragraph(f"<b>Tổng thu nhập:</b> {total_income:,.0f} VNĐ", summary_style))
            elements.append(Paragraph(f"<b>Tổng chi tiêu:</b> {total_expense:,.0f} VNĐ", summary_style))
            elements.append(Paragraph(f"<b>Số dư:</b> {balance:,.0f} VNĐ", summary_style))

            elements.append(Spacer(1, 0.3*inch))

            # Thời gian xuất
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=TA_CENTER,
                fontName=font_name
            )
            elements.append(Paragraph(
                f"Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                footer_style
            ))

            # Tạo PDF
            doc.build(elements)

            messagebox.showinfo("Thành công", f"Đã xuất file PDF thành công!\n{filename}")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất PDF: {str(e)}")

    def __del__(self):
        """Đóng kết nối database khi thoát"""
        if hasattr(self, 'conn'):
            self.conn.close()

    def import_from_excel(self):
        """Đọc và nhập dữ liệu giao dịch từ file Excel"""
        try:
            import pandas as pd
        except ImportError:
            messagebox.showerror("Lỗi", "Vui lòng cài đặt thư viện pandas và openpyxl:\npip install pandas openpyxl")
            return

        # 1. Chọn file Excel
        filepath = filedialog.askopenfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
            title="Chọn file Excel để nhập dữ liệu"
        )

        if not filepath:
            return  # Người dùng hủy

        try:
            # 2. Đọc dữ liệu từ file Excel (sheet đầu tiên)
            # Giả định cột: Loại, Danh mục, Số tiền, Mô tả, Ngày
            df = pd.read_excel(filepath)

            # *** BỔ SUNG: Làm sạch tên cột (loại bỏ khoảng trắng thừa) ***
            df.columns = df.columns.str.strip()

            # Đảm bảo các cột cần thiết tồn tại và chuẩn hóa tên cột
            required_columns = ['Loại', 'Danh mục', 'Số tiền', 'Mô tả', 'Ngày']
            if not all(col in df.columns for col in required_columns):
                messagebox.showerror("Lỗi", f"File Excel phải có đủ các cột: {', '.join(required_columns)}")
                return

            # Chuyển đổi dữ liệu và chuẩn hóa
            df = df[required_columns].copy()
            df.columns = ['type', 'category', 'amount', 'description', 'date']

            # Chuẩn hóa loại giao dịch (income/expense)
            df['type'] = df['type'].str.lower().replace({
                'thu nhập': 'income', 'chi tiêu': 'expense',
                'thu': 'income', 'chi': 'expense'
            })

            # Lọc các dòng không hợp lệ
            df = df[df['type'].isin(['income', 'expense'])]
            df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
            df['date'] = pd.to_datetime(df['date'], errors='coerce')

            # Loại bỏ các dòng bị lỗi (Số tiền không hợp lệ, Ngày không hợp lệ)
            df.dropna(subset=['amount', 'date'], inplace=True)
            df = df[df['amount'] > 0]

            # Định dạng lại ngày tháng theo chuẩn YYYY-MM-DD
            df['date'] = df['date'].dt.strftime('%Y-%m-%d')

            # Lấy danh sách danh mục hiện có để kiểm tra
            self.cursor.execute('SELECT name, type FROM categories')
            existing_categories = {(name, type) for name, type in self.cursor.fetchall()}

            new_transactions = []

            # Chuẩn bị dữ liệu để chèn
            for index, row in df.iterrows():
                trans_type = row['type']
                category = str(row['category']).strip() if row['category'] else 'Khác'

                # Kiểm tra và thêm danh mục mới nếu cần
                if (category, trans_type) not in existing_categories and category != 'Khác':
                    self.cursor.execute('INSERT INTO categories (name, type) VALUES (?, ?)', (category, trans_type))
                    existing_categories.add((category, trans_type))

                new_transactions.append((
                    trans_type,
                    category,
                    row['amount'],
                    row['description'] if row['description'] else '',
                    row['date'],
                    self.user_id
                ))

            if not new_transactions:
                messagebox.showwarning("Cảnh báo", "Không tìm thấy giao dịch hợp lệ nào trong file Excel.")
                return

            # 3. Chèn dữ liệu vào database
            self.cursor.executemany('''
                INSERT INTO transactions (type, category, amount, description, date, user_id)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', new_transactions)

            self.conn.commit()
            messagebox.showinfo("Thành công", f"Đã nhập thành công {len(new_transactions)} giao dịch từ Excel!")

            # Cập nhật danh sách và thống kê
            self.update_categories()  # Cập nhật danh mục mới (nếu có)
            self.load_transactions()

        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi đọc file Excel: {e}")
            # print(e) # In lỗi ra console để debug nếu cần

    def export_to_excel(self):
        """Xuất danh sách giao dịch ra file Excel"""
        try:
            import pandas as pd
        except ImportError:
            messagebox.showerror("Lỗi", "Vui lòng cài đặt thư viện pandas và openpyxl:\npip install pandas openpyxl")
            return

        # Lấy bộ lọc hiện tại
        filter_month = self.filter_month_var.get()
        filter_year = self.filter_year_var.get()

        # Truy vấn dữ liệu
        query = 'SELECT * FROM transactions WHERE user_id = ?'
        params = [self.user_id]

        if filter_month != "Tất cả":
            query += ' AND strftime("%m", date) = ?'
            params.append(filter_month.zfill(2))

        if filter_year != "Tất cả":
            query += ' AND strftime("%Y", date) = ?'
            params.append(filter_year)

        query += ' ORDER BY date DESC'

        self.cursor.execute(query, params)
        transactions = self.cursor.fetchall()

        if not transactions:
            messagebox.showinfo("Thông báo", "Không có dữ liệu để xuất!")
            return

        # Chọn nơi lưu file
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"giao_dich_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not filename:
            return

        try:
            # Chuẩn bị dữ liệu cho DataFrame
            data = []
            total_income = 0
            total_expense = 0

            for trans in transactions:
                trans_id, trans_type, category, amount, description, date, user_id = trans
                type_text = "Thu nhập" if trans_type == "income" else "Chi tiêu"

                # Đổi định dạng ngày
                date_parts = date.split('-')
                date_formatted = f"{date_parts[2]}/{date_parts[1]}/{date_parts[0]}"

                data.append({
                    'ID': trans_id,
                    'Loại': type_text,
                    'Danh mục': category,
                    'Số tiền': amount,
                    'Mô tả': description or '',
                    'Ngày': date_formatted
                })

                if trans_type == "income":
                    total_income += amount
                else:
                    total_expense += amount

            # Tạo DataFrame
            df = pd.DataFrame(data)

            # Tạo writer với engine openpyxl
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Xuất dữ liệu giao dịch
                df.to_excel(writer, sheet_name='Giao dịch', index=False)

                # Lấy workbook và worksheet để format
                workbook = writer.book
                worksheet = writer.sheets['Giao dịch']

                # Định dạng header
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Định dạng cột số tiền
                for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=4, max_col=4):
                    for cell in row:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right')

                # Căn giữa các cột
                for col in ['A', 'B', 'F']:  # ID, Loại, Ngày
                    for cell in worksheet[col]:
                        cell.alignment = Alignment(horizontal='center')

                # Điều chỉnh độ rộng cột
                worksheet.column_dimensions['A'].width = 8   # ID
                worksheet.column_dimensions['B'].width = 12  # Loại
                worksheet.column_dimensions['C'].width = 15  # Danh mục
                worksheet.column_dimensions['D'].width = 15  # Số tiền
                worksheet.column_dimensions['E'].width = 30  # Mô tả
                worksheet.column_dimensions['F'].width = 12  # Ngày

                # Thêm sheet thống kê
                balance = total_income - total_expense
                summary_data = {
                    'Chỉ số': ['Tổng thu nhập', 'Tổng chi tiêu', 'Số dư'],
                    'Số tiền (VNĐ)': [total_income, total_expense, balance]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Thống kê', index=False)

                # Format sheet thống kê
                worksheet_summary = writer.sheets['Thống kê']
                
                for cell in worksheet_summary[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                for row in worksheet_summary.iter_rows(min_row=2, max_row=4, min_col=2, max_col=2):
                    for cell in row:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right')

                worksheet_summary.column_dimensions['A'].width = 20
                worksheet_summary.column_dimensions['B'].width = 20

                # Tô màu cho các dòng thống kê
                green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                blue_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")

                for cell in worksheet_summary[2]:  # Thu nhập
                    cell.fill = green_fill
                for cell in worksheet_summary[3]:  # Chi tiêu
                    cell.fill = red_fill
                for cell in worksheet_summary[4]:  # Số dư
                    cell.fill = blue_fill

            messagebox.showinfo("Thành công", f"Đã xuất {len(df)} giao dịch ra file Excel!\n{filename}")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất Excel: {str(e)}")

    def show_monthly_chart(self):
        """Hiển thị biểu đồ theo tháng (có thêm phần trăm)"""
        filter_year = self.filter_year_var.get()

        if filter_year == "Tất cả":
            filter_year = str(datetime.now().year)

        # Lấy dữ liệu thu nhập và chi tiêu theo tháng
        months = list(range(1, 13))
        income_data = []
        expense_data = []

        for month in months:
            month_str = str(month).zfill(2)
            # Thu nhập
            self.cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) FROM transactions 
                WHERE type = "income" AND strftime("%m", date) = ? AND strftime("%Y", date) = ?
            ''', (month_str, filter_year))
            income_data.append(self.cursor.fetchone()[0])

            # Chi tiêu
            self.cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) FROM transactions 
                WHERE type = "expense" AND strftime("%m", date) = ? AND strftime("%Y", date) = ?
            ''', (month_str, filter_year))
            expense_data.append(self.cursor.fetchone()[0])

        # Tạo cửa sổ mới
        chart_window = tk.Toplevel(self.root)
        chart_window.title(f"Biểu Đồ Tài Chính Năm {filter_year}")
        chart_window.geometry("900x600")

        fig = Figure(figsize=(9, 6))
        ax = fig.add_subplot(111)

        x = range(1, 13)
        width = 0.35

        # Vẽ Bar
        rects_income = ax.bar([i - width / 2 for i in x], income_data, width, label='Thu nhập', color='#4CAF50')
        rects_expense = ax.bar([i + width / 2 for i in x], expense_data, width, label='Chi tiêu', color='#f44336')

        ax.set_xlabel('Tháng', fontsize=12)
        ax.set_ylabel('Số tiền (VNĐ)', fontsize=12)
        ax.set_title(f'Biểu Đồ Tài Chính Theo Tháng - Năm {filter_year}',
                     fontsize=14, fontweight='bold')
        ax.set_xticks(x)
        ax.legend()
        ax.grid(axis='y', alpha=0.3)

        # *** THÊM PHẦN TRĂM ***
        def autolabel(rects, data_list):
            for i, rect in enumerate(rects):
                height = rect.get_height()
                # Tổng thu và chi trong tháng
                total = income_data[i] + expense_data[i]

                # Chỉ hiển thị phần trăm nếu total > 0 (tránh chia cho 0)
                if total > 0:
                    percentage = (data_list[i] / total) * 100
                    # Vị trí đặt nhãn: trên cùng của cột, dịch sang phải/trái một chút
                    x_pos = rect.get_x() + rect.get_width() / 2
                    y_pos = height
                    ax.text(x_pos, y_pos, f'{percentage:.0f}%',
                            ha='center', va='bottom', fontsize=8, color='black',
                            rotation=45)  # Xoay 45 độ để tránh chồng chéo

        autolabel(rects_income, income_data)
        autolabel(rects_expense, expense_data)
        # *** HẾT PHẦN THÊM PHẦN TRĂM ***

        canvas = FigureCanvasTkAgg(fig, master=chart_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def show_yearly_chart(self):
        """Hiển thị biểu đồ theo năm (có thêm phần trăm)"""
        # Lấy danh sách các năm có dữ liệu
        self.cursor.execute('''
            SELECT DISTINCT strftime("%Y", date) FROM transactions 
            ORDER BY date
        ''')
        years = [row[0] for row in self.cursor.fetchall()]

        if not years:
            messagebox.showinfo("Thông báo", "Không có dữ liệu để hiển thị!")
            return

        income_data = []
        expense_data = []

        for year in years:
            # Thu nhập
            self.cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) FROM transactions 
                WHERE type = "income" AND strftime("%Y", date) = ?
            ''', (year,))
            income_data.append(self.cursor.fetchone()[0])

            # Chi tiêu
            self.cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) FROM transactions 
                WHERE type = "expense" AND strftime("%Y", date) = ?
            ''', (year,))
            expense_data.append(self.cursor.fetchone()[0])

        # Tạo cửa sổ mới
        chart_window = tk.Toplevel(self.root)
        chart_window.title("Biểu Đồ Tài Chính Theo Năm")
        chart_window.geometry("900x600")

        fig = Figure(figsize=(9, 6))
        ax = fig.add_subplot(111)

        x = range(len(years))
        width = 0.35

        # Vẽ Bar
        rects_income = ax.bar([i - width / 2 for i in x], income_data, width, label='Thu nhập', color='#4CAF50')
        rects_expense = ax.bar([i + width / 2 for i in x], expense_data, width, label='Chi tiêu', color='#f44336')

        ax.set_xlabel('Năm', fontsize=12)
        ax.set_ylabel('Số tiền (VNĐ)', fontsize=12)
        ax.set_title('Biểu Đồ Tài Chính Theo Năm', fontsize=14, fontweight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels(years)
        ax.legend()
        ax.grid(axis='y', alpha=0.3)

        # *** THÊM PHẦN TRĂM ***
        def autolabel_yearly(rects, data_list):
            for i, rect in enumerate(rects):
                height = rect.get_height()
                # Tổng thu và chi trong năm
                total = income_data[i] + expense_data[i]

                # Chỉ hiển thị phần trăm nếu total > 0
                if total > 0:
                    percentage = (data_list[i] / total) * 100
                    x_pos = rect.get_x() + rect.get_width() / 2
                    y_pos = height
                    ax.text(x_pos, y_pos, f'{percentage:.0f}%',
                            ha='center', va='bottom', fontsize=9, color='black')

        autolabel_yearly(rects_income, income_data)
        autolabel_yearly(rects_expense, expense_data)
        # *** HẾT PHẦN THÊM PHẦN TRĂM ***

        canvas = FigureCanvasTkAgg(fig, master=chart_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def set_budget_limit(self):
        """Đặt hạn mức chi tiêu hàng tháng"""
        # Tạo cửa sổ đặt hạn mức
        budget_window = tk.Toplevel(self.root)
        budget_window.title("Đặt Hạn Mức Chi Tiêu")
        budget_window.geometry("500x400")
        budget_window.configure(bg="#f0f0f0")

        # Frame chính
        main_frame = tk.LabelFrame(budget_window, text="Thiết Lập Hạn Mức Chi Tiêu Hàng Tháng",
                                   bg="white", font=("Arial", 12, "bold"),
                                   padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Chọn tháng
        tk.Label(main_frame, text="Chọn Tháng:", bg="white",
                font=("Arial", 10)).grid(row=0, column=0, sticky="w", pady=10)
        
        month_var = tk.StringVar(value=str(datetime.now().month))
        month_combo = ttk.Combobox(main_frame, textvariable=month_var, state="readonly", width=20)
        month_combo['values'] = [str(i) for i in range(1, 13)]
        month_combo.grid(row=0, column=1, sticky="w", pady=10, padx=10)

        # Chọn năm
        tk.Label(main_frame, text="Chọn Năm:", bg="white",
                font=("Arial", 10)).grid(row=1, column=0, sticky="w", pady=10)
        
        year_var = tk.StringVar(value=str(datetime.now().year))
        year_combo = ttk.Combobox(main_frame, textvariable=year_var, state="readonly", width=20)
        year_combo['values'] = [str(i) for i in range(2020, 2031)]
        year_combo.grid(row=1, column=1, sticky="w", pady=10, padx=10)

        # Nhập hạn mức
        tk.Label(main_frame, text="Hạn Mức (VNĐ):", bg="white",
                font=("Arial", 10)).grid(row=2, column=0, sticky="w", pady=10)
        
        limit_var = tk.StringVar()
        limit_entry = tk.Entry(main_frame, textvariable=limit_var, width=22,
                              font=("Arial", 10))
        limit_entry.grid(row=2, column=1, sticky="w", pady=10, padx=10)

        # Hiển thị hạn mức hiện tại
        current_limit_label = tk.Label(main_frame, text="", bg="white",
                                       font=("Arial", 9), fg="#666")
        current_limit_label.grid(row=3, column=0, columnspan=2, pady=10)

        def load_current_limit(*args):
            """Load hạn mức hiện tại khi thay đổi tháng/năm"""
            month = int(month_var.get())
            year = int(year_var.get())
            
            self.cursor.execute('''
                SELECT limit_amount FROM budget_limits 
                WHERE user_id = ? AND month = ? AND year = ?
            ''', (self.user_id, month, year))
            
            result = self.cursor.fetchone()
            if result:
                current_limit_label.config(
                    text=f"Hạn mức hiện tại: {result[0]:,.0f} VNĐ",
                    fg="#4CAF50"
                )
                limit_var.set(str(int(result[0])))
            else:
                current_limit_label.config(
                    text="Chưa đặt hạn mức cho tháng này",
                    fg="#999"
                )
                limit_var.set("")

        # Gắn sự kiện thay đổi tháng/năm
        month_var.trace('w', load_current_limit)
        year_var.trace('w', load_current_limit)
        
        # Load hạn mức ban đầu
        load_current_limit()

        def save_budget_limit():
            """Lưu hạn mức chi tiêu"""
            try:
                month = int(month_var.get())
                year = int(year_var.get())
                limit_amount = float(limit_var.get().replace(',', ''))

                if limit_amount <= 0:
                    messagebox.showerror("Lỗi", "Hạn mức phải lớn hơn 0!")
                    return

                # Kiểm tra xem đã có hạn mức chưa
                self.cursor.execute('''
                    SELECT id FROM budget_limits 
                    WHERE user_id = ? AND month = ? AND year = ?
                ''', (self.user_id, month, year))
                
                existing = self.cursor.fetchone()
                
                if existing:
                    # Cập nhật hạn mức
                    self.cursor.execute('''
                        UPDATE budget_limits 
                        SET limit_amount = ?
                        WHERE user_id = ? AND month = ? AND year = ?
                    ''', (limit_amount, self.user_id, month, year))
                    message = f"Đã cập nhật hạn mức tháng {month}/{year}: {limit_amount:,.0f} VNĐ"
                else:
                    # Thêm hạn mức mới
                    self.cursor.execute('''
                        INSERT INTO budget_limits (user_id, month, year, limit_amount)
                        VALUES (?, ?, ?, ?)
                    ''', (self.user_id, month, year, limit_amount))
                    message = f"Đã đặt hạn mức tháng {month}/{year}: {limit_amount:,.0f} VNĐ"
                
                self.conn.commit()
                messagebox.showinfo("Thành công", message)
                load_current_limit()
                self.update_budget_info_display()  # Cập nhật bảng thông báo
                self.check_budget_warning()  # Kiểm tra cảnh báo sau khi đặt hạn mức
                
            except ValueError:
                messagebox.showerror("Lỗi", "Vui lòng nhập số tiền hợp lệ!")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể lưu hạn mức: {str(e)}")

        def delete_budget_limit():
            """Xóa hạn mức chi tiêu"""
            month = int(month_var.get())
            year = int(year_var.get())
            
            if messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa hạn mức tháng {month}/{year}?"):
                self.cursor.execute('''
                    DELETE FROM budget_limits 
                    WHERE user_id = ? AND month = ? AND year = ?
                ''', (self.user_id, month, year))
                
                self.conn.commit()
                messagebox.showinfo("Thành công", f"Đã xóa hạn mức tháng {month}/{year}")
                load_current_limit()
                self.update_budget_info_display()  # Cập nhật bảng thông báo

        # Nút lưu và xóa
        button_frame = tk.Frame(main_frame, bg="white")
        button_frame.grid(row=4, column=0, columnspan=2, pady=20)

        tk.Button(button_frame, text="💾 Lưu Hạn Mức", command=save_budget_limit,
                 bg="#4CAF50", fg="white", font=("Arial", 10, "bold"),
                 padx=20, pady=8, cursor="hand2").pack(side=tk.LEFT, padx=5)

        tk.Button(button_frame, text="🗑️ Xóa Hạn Mức", command=delete_budget_limit,
                 bg="#f44336", fg="white", font=("Arial", 10, "bold"),
                 padx=20, pady=8, cursor="hand2").pack(side=tk.LEFT, padx=5)

        # Hiển thị chi tiêu hiện tại của tháng
        expense_info_label = tk.Label(main_frame, text="", bg="white",
                                     font=("Arial", 10, "bold"), fg="#333")
        expense_info_label.grid(row=5, column=0, columnspan=2, pady=10)

        def update_expense_info(*args):
            """Cập nhật thông tin chi tiêu hiện tại"""
            month = int(month_var.get())
            year = int(year_var.get())
            
            month_str = str(month).zfill(2)
            year_str = str(year)
            
            self.cursor.execute('''
                SELECT COALESCE(SUM(amount), 0) FROM transactions 
                WHERE user_id = ? AND type = "expense" 
                AND strftime("%m", date) = ? AND strftime("%Y", date) = ?
            ''', (self.user_id, month_str, year_str))
            
            total_expense = self.cursor.fetchone()[0]
            
            # Lấy hạn mức
            self.cursor.execute('''
                SELECT limit_amount FROM budget_limits 
                WHERE user_id = ? AND month = ? AND year = ?
            ''', (self.user_id, month, year))
            
            result = self.cursor.fetchone()
            
            info_text = f"Chi tiêu hiện tại tháng {month}/{year}: {total_expense:,.0f} VNĐ"
            
            if result:
                limit_amount = result[0]
                remaining = limit_amount - total_expense
                percentage = (total_expense / limit_amount * 100) if limit_amount > 0 else 0
                
                info_text += f"\nCòn lại: {remaining:,.0f} VNĐ ({100-percentage:.1f}%)"
                
                if total_expense > limit_amount:
                    expense_info_label.config(fg="#f44336")  # Đỏ - vượt hạn mức
                elif percentage >= 80:
                    expense_info_label.config(fg="#FF9800")  # Cam - gần hạn mức
                else:
                    expense_info_label.config(fg="#4CAF50")  # Xanh - OK
            else:
                expense_info_label.config(fg="#333")
            
            expense_info_label.config(text=info_text)

        month_var.trace('w', update_expense_info)
        year_var.trace('w', update_expense_info)
        update_expense_info()

    def manage_categories(self):
        """Quản lý danh mục (Thêm, Sửa, Xóa)"""
        # Tạo cửa sổ quản lý danh mục
        category_window = tk.Toplevel(self.root)
        category_window.title("Quản Lý Danh Mục")
        category_window.geometry("700x500")
        category_window.configure(bg="#f0f0f0")

        # Frame chính
        main_frame = tk.Frame(category_window, bg="#f0f0f0", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Frame trái - Thêm/Sửa danh mục
        left_frame = tk.LabelFrame(main_frame, text="Thêm/Sửa Danh Mục",
                                   bg="white", font=("Arial", 11, "bold"),
                                   padx=15, pady=15)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # Tên danh mục
        tk.Label(left_frame, text="Tên danh mục:", bg="white", font=("Arial", 10)).grid(
            row=0, column=0, sticky="w", pady=10)
        category_name_var = tk.StringVar()
        category_name_entry = tk.Entry(left_frame, textvariable=category_name_var,
                                       width=25, font=("Arial", 10))
        category_name_entry.grid(row=0, column=1, sticky="w", pady=10)

        # Loại danh mục
        tk.Label(left_frame, text="Loại:", bg="white", font=("Arial", 10)).grid(
            row=1, column=0, sticky="w", pady=10)
        category_type_var = tk.StringVar(value="expense")
        type_frame = tk.Frame(left_frame, bg="white")
        type_frame.grid(row=1, column=1, sticky="w", pady=10)
        tk.Radiobutton(type_frame, text="Thu nhập", variable=category_type_var,
                      value="income", bg="white").pack(side=tk.LEFT, padx=5)
        tk.Radiobutton(type_frame, text="Chi tiêu", variable=category_type_var,
                      value="expense", bg="white").pack(side=tk.LEFT, padx=5)

        # Biến lưu ID danh mục đang sửa
        editing_category_id = [None]  # Dùng list để có thể thay đổi trong nested function

        # Nút thêm
        def add_category():
            name = category_name_var.get().strip()
            cat_type = category_type_var.get()

            if not name:
                messagebox.showerror("Lỗi", "Vui lòng nhập tên danh mục!")
                return

            try:
                # Kiểm tra trùng tên
                self.cursor.execute('SELECT COUNT(*) FROM categories WHERE name = ? AND type = ?',
                                   (name, cat_type))
                if self.cursor.fetchone()[0] > 0:
                    messagebox.showerror("Lỗi", "Danh mục này đã tồn tại!")
                    return

                self.cursor.execute('INSERT INTO categories (name, type) VALUES (?, ?)',
                                   (name, cat_type))
                self.conn.commit()
                messagebox.showinfo("Thành công", "Đã thêm danh mục mới!")

                # Reset form
                category_name_var.set("")
                editing_category_id[0] = None

                # Cập nhật danh sách
                load_categories_list()
                self.update_categories()
                self.update_filter_categories()

            except Exception as e:
                messagebox.showerror("Lỗi", f"Có lỗi xảy ra: {str(e)}")

        # Nút sửa
        def update_category():
            if editing_category_id[0] is None:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn danh mục để sửa!")
                return

            name = category_name_var.get().strip()
            cat_type = category_type_var.get()

            if not name:
                messagebox.showerror("Lỗi", "Vui lòng nhập tên danh mục!")
                return

            try:
                # Kiểm tra trùng tên (trừ chính nó)
                self.cursor.execute('''SELECT COUNT(*) FROM categories 
                                      WHERE name = ? AND type = ? AND id != ?''',
                                   (name, cat_type, editing_category_id[0]))
                if self.cursor.fetchone()[0] > 0:
                    messagebox.showerror("Lỗi", "Tên danh mục này đã tồn tại!")
                    return

                # Lấy tên cũ
                self.cursor.execute('SELECT name, type FROM categories WHERE id = ?',
                                   (editing_category_id[0],))
                old_name, old_type = self.cursor.fetchone()

                # Cập nhật danh mục
                self.cursor.execute('UPDATE categories SET name = ?, type = ? WHERE id = ?',
                                   (name, cat_type, editing_category_id[0]))

                # Cập nhật các giao dịch sử dụng danh mục này
                self.cursor.execute('UPDATE transactions SET category = ?, type = ? WHERE category = ? AND type = ?',
                                   (name, cat_type, old_name, old_type))

                self.conn.commit()
                messagebox.showinfo("Thành công", "Đã cập nhật danh mục!")

                # Reset form
                category_name_var.set("")
                editing_category_id[0] = None

                # Cập nhật danh sách
                load_categories_list()
                self.update_categories()
                self.update_filter_categories()
                self.load_transactions()

            except Exception as e:
                messagebox.showerror("Lỗi", f"Có lỗi xảy ra: {str(e)}")

        # Nút hủy sửa
        def cancel_edit():
            category_name_var.set("")
            editing_category_id[0] = None
            add_btn.config(state=tk.NORMAL)
            update_btn.config(state=tk.DISABLED)
            cancel_btn.config(state=tk.DISABLED)

        # Các nút
        btn_frame = tk.Frame(left_frame, bg="white")
        btn_frame.grid(row=2, column=0, columnspan=2, pady=20)

        add_btn = tk.Button(btn_frame, text="Thêm Danh Mục", command=add_category,
                           bg="#4CAF50", fg="white", font=("Arial", 10, "bold"),
                           cursor="hand2", padx=15, pady=8)
        add_btn.pack(side=tk.LEFT, padx=5)

        update_btn = tk.Button(btn_frame, text="Cập Nhật", command=update_category,
                              bg="#FF9800", fg="white", font=("Arial", 10, "bold"),
                              cursor="hand2", padx=15, pady=8, state=tk.DISABLED)
        update_btn.pack(side=tk.LEFT, padx=5)

        cancel_btn = tk.Button(btn_frame, text="Hủy", command=cancel_edit,
                              bg="#9E9E9E", fg="white", font=("Arial", 10, "bold"),
                              cursor="hand2", padx=15, pady=8, state=tk.DISABLED)
        cancel_btn.pack(side=tk.LEFT, padx=5)

        # Frame phải - Danh sách danh mục
        right_frame = tk.LabelFrame(main_frame, text="Danh Sách Danh Mục",
                                    bg="white", font=("Arial", 11, "bold"),
                                    padx=15, pady=15)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Bảng danh mục
        tree_frame = tk.Frame(right_frame, bg="white")
        tree_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("ID", "Tên", "Loại")
        category_tree = ttk.Treeview(tree_frame, columns=columns,
                                     show="headings", height=15)

        category_tree.heading("ID", text="ID")
        category_tree.heading("Tên", text="Tên Danh Mục")
        category_tree.heading("Loại", text="Loại")

        category_tree.column("ID", width=50, anchor="center")
        category_tree.column("Tên", width=150)
        category_tree.column("Loại", width=100, anchor="center")

        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL,
                                 command=category_tree.yview)
        category_tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        category_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Load danh sách danh mục
        def load_categories_list():
            for item in category_tree.get_children():
                category_tree.delete(item)

            self.cursor.execute('SELECT id, name, type FROM categories ORDER BY type, name')
            categories = self.cursor.fetchall()

            for cat in categories:
                cat_id, name, cat_type = cat
                type_text = "Thu nhập" if cat_type == "income" else "Chi tiêu"
                category_tree.insert("", tk.END, values=(cat_id, name, type_text))

        load_categories_list()

        # Sự kiện chọn danh mục để sửa
        def on_category_select(event):
            selected = category_tree.selection()
            if selected:
                item = category_tree.item(selected[0])
                cat_id, name, type_text = item['values']

                category_name_var.set(name)
                category_type_var.set("income" if type_text == "Thu nhập" else "expense")
                editing_category_id[0] = cat_id

                add_btn.config(state=tk.DISABLED)
                update_btn.config(state=tk.NORMAL)
                cancel_btn.config(state=tk.NORMAL)

        category_tree.bind('<<TreeviewSelect>>', on_category_select)

        # Nút xóa danh mục
        def delete_category():
            selected = category_tree.selection()
            if not selected:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn danh mục để xóa!")
                return

            item = category_tree.item(selected[0])
            cat_id, name, type_text = item['values']
            cat_type = "income" if type_text == "Thu nhập" else "expense"

            # Kiểm tra xem danh mục có đang được sử dụng không
            self.cursor.execute('SELECT COUNT(*) FROM transactions WHERE category = ? AND type = ?',
                               (name, cat_type))
            count = self.cursor.fetchone()[0]

            if count > 0:
                if not messagebox.askyesno("Xác nhận",
                    f"Danh mục '{name}' đang được sử dụng trong {count} giao dịch.\n"
                    f"Nếu xóa, các giao dịch này sẽ chuyển sang danh mục 'Khác'.\n"
                    f"Bạn có chắc muốn tiếp tục?"):
                    return

                # Chuyển các giao dịch sang danh mục "Khác"
                self.cursor.execute('''UPDATE transactions SET category = 'Khác' 
                                      WHERE category = ? AND type = ?''',
                                   (name, cat_type))

            # Xóa danh mục
            if messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa danh mục '{name}'?"):
                self.cursor.execute('DELETE FROM categories WHERE id = ?', (cat_id,))
                self.conn.commit()
                messagebox.showinfo("Thành công", "Đã xóa danh mục!")

                # Reset form và cập nhật
                cancel_edit()
                load_categories_list()
                self.update_categories()
                self.update_filter_categories()
                self.load_transactions()

        delete_btn_frame = tk.Frame(right_frame, bg="white")
        delete_btn_frame.pack(pady=10)

        tk.Button(delete_btn_frame, text="Xóa Danh Mục", command=delete_category,
                 bg="#f44336", fg="white", font=("Arial", 10, "bold"),
                 cursor="hand2", padx=20, pady=8).pack()

    def show_ai_config_help(self):
        """Hiển thị hướng dẫn cấu hình AI"""
        help_window = tk.Toplevel(self.root)
        help_window.title("Cấu hình ChatBot AI")
        help_window.geometry("600x400")
        help_window.configure(bg="white")
        
        # Frame chính
        main_frame = tk.Frame(help_window, bg="white", padx=30, pady=30)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Tiêu đề
        tk.Label(main_frame, text="⚙️ Cấu hình ChatBot AI", 
                bg="white", fg="#4285F4",
                font=("Arial", 16, "bold")).pack(pady=(0, 20))
        
        # Nội dung hướng dẫn
        help_text = """
🤖 Để sử dụng ChatBot AI, bạn cần:

1️⃣ Cài đặt thư viện Google Generative AI:
   • Mở Terminal/Command Prompt
   • Chạy lệnh: pip install google-generativeai

2️⃣ Lấy API Key miễn phí từ Google:
   • Truy cập: https://makersuite.google.com/app/apikey
   • Đăng nhập bằng tài khoản Google
   • Click "Create API Key"
   • Copy API Key

3️⃣ Cấu hình API Key:
   • Mở file config.py trong thư mục ứng dụng
   • Dán API Key vào giữa dấu ngoặc kép
   • Lưu file lại

4️⃣ Khởi động lại ứng dụng

✨ Sau khi cấu hình xong, bạn có thể sử dụng ChatBot AI 
để phân tích chi tiêu và nhận lời khuyên tài chính!
"""
        
        text_widget = tk.Text(main_frame, wrap=tk.WORD, 
                             font=("Arial", 10),
                             bg="#f9f9f9", relief=tk.FLAT,
                             padx=15, pady=15, height=15)
        text_widget.pack(fill=tk.BOTH, expand=True)
        text_widget.insert(1.0, help_text)
        text_widget.config(state=tk.DISABLED)
        
        # Nút đóng
        tk.Button(main_frame, text="Đóng",
                 command=help_window.destroy,
                 bg="#4285F4", fg="white",
                 font=("Arial", 10, "bold"),
                 cursor="hand2", padx=30, pady=8).pack(pady=(15, 0))

    def open_chatbot(self):
        """Mở cửa sổ ChatBot AI"""
        # Kiểm tra xem ChatBot có sẵn không
        if not CHATBOT_AVAILABLE:
            messagebox.showwarning(
                "Chưa cài đặt thư viện", 
                "Vui lòng cài đặt thư viện Google Generative AI:\n\n"
                "pip install google-generativeai\n\n"
                "Sau đó khởi động lại ứng dụng."
            )
            self.show_ai_config_help()
            return
        
        if not self.chatbot:
            messagebox.showwarning(
                "Lỗi khởi tạo ChatBot", 
                "Không thể khởi tạo ChatBot.\n"
                "Vui lòng kiểm tra cấu hình API Key."
            )
            self.show_ai_config_help()
            return
            
        if not self.chatbot.is_available():
            messagebox.showinfo(
                "Chưa cấu hình API Key", 
                "Bạn chưa nhập API Key cho ChatBot AI.\n\n"
                "Vui lòng:\n"
                "1. Lấy API Key miễn phí tại:\n"
                "   https://makersuite.google.com/app/apikey\n\n"
                "2. Mở file config.py và nhập API Key\n\n"
                "3. Khởi động lại ứng dụng"
            )
            self.show_ai_config_help()
            return
        
        chatbot_window = tk.Toplevel(self.root)
        chatbot_window.title("🤖 Trợ Lý Tài Chính AI - Google Gemini")
        chatbot_window.geometry("750x650")
        chatbot_window.configure(bg="#f5f5f5")
        
        # Frame chính
        main_frame = tk.Frame(chatbot_window, bg="white", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Tiêu đề
        title_frame = tk.Frame(main_frame, bg="#4285F4", pady=12)
        title_frame.pack(fill=tk.X)
        
        tk.Label(title_frame, text="🤖 Trợ Lý Tài Chính AI",
                bg="#4285F4", fg="white",
                font=("Arial", 14, "bold")).pack()
        
        tk.Label(title_frame, text="Powered by Google Gemini",
                bg="#4285F4", fg="white",
                font=("Arial", 8)).pack()
        
        # Khu vực hiển thị chat
        chat_frame = tk.Frame(main_frame, bg="white")
        chat_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Text widget để hiển thị hội thoại
        self.chat_display = tk.Text(chat_frame, wrap=tk.WORD, 
                                    font=("Arial", 10),
                                    bg="#f9f9f9", relief=tk.FLAT,
                                    padx=10, pady=10, state=tk.NORMAL)
        self.chat_display.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Scrollbar
        scrollbar = tk.Scrollbar(chat_frame, command=self.chat_display.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.chat_display.config(yscrollcommand=scrollbar.set)
        
        # Tag để format tin nhắn
        self.chat_display.tag_config("user", foreground="#1976D2", 
                                    font=("Arial", 10, "bold"))
        self.chat_display.tag_config("bot", foreground="#34A853", 
                                    font=("Arial", 10))
        self.chat_display.tag_config("system", foreground="#666", 
                                    font=("Arial", 9, "italic"))
        self.chat_display.tag_config("time", foreground="#999", 
                                    font=("Arial", 8))
        
        # Frame nhập tin nhắn
        input_frame = tk.Frame(main_frame, bg="white")
        input_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Entry nhập tin nhắn
        self.message_var = tk.StringVar()
        message_entry = tk.Entry(input_frame, textvariable=self.message_var,
                                font=("Arial", 11), relief=tk.FLAT,
                                bg="#f0f0f0", bd=2)
        message_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10), ipady=5)
        message_entry.bind('<Return>', lambda e: self.send_message())
        message_entry.focus()
        
        # Nút gửi
        send_btn = tk.Button(input_frame, text="Gửi ➤",
                            command=self.send_message,
                            bg="#34A853", fg="white",
                            font=("Arial", 10, "bold"),
                            cursor="hand2", padx=20, relief=tk.FLAT)
        send_btn.pack(side=tk.LEFT)
        
        # Frame các nút gợi ý
        suggestion_frame = tk.LabelFrame(main_frame, text="💡 Câu hỏi gợi ý",
                                        bg="white", font=("Arial", 9, "bold"),
                                        fg="#4285F4")
        suggestion_frame.pack(fill=tk.X, pady=(10, 0))
        
        suggestions = [
            ("💡", "Cho tôi lời khuyên tài chính"),
            ("📊", "Phân tích xu hướng chi tiêu"),
            ("💰", "Làm sao để tiết kiệm hiệu quả?"),
            ("🎯", "Tôi nên đặt ngân sách như thế nào?"),
            ("📈", "Đánh giá tình hình tài chính của tôi"),
            ("🔍", "Tìm cách giảm chi tiêu không cần thiết")
        ]
        
        # Tạo 3 hàng, mỗi hàng 2 nút
        for row in range(3):
            btn_row = tk.Frame(suggestion_frame, bg="white")
            btn_row.pack(fill=tk.X, pady=2)
            
            for col in range(2):
                idx = row * 2 + col
                if idx < len(suggestions):
                    emoji, text = suggestions[idx]
                    tk.Button(btn_row, text=f"{emoji} {text}",
                             command=lambda t=text: self.send_suggestion(t),
                             bg="#E8F0FE", fg="#1967D2",
                             font=("Arial", 8),
                             cursor="hand2", relief=tk.FLAT,
                             padx=8, pady=6).pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)
        
        # Nút reset chat
        reset_frame = tk.Frame(main_frame, bg="white")
        reset_frame.pack(fill=tk.X, pady=(5, 0))
        
        tk.Button(reset_frame, text="🔄 Bắt đầu cuộc trò chuyện mới",
                 command=self.reset_chat,
                 bg="#9E9E9E", fg="white",
                 font=("Arial", 8),
                 cursor="hand2", relief=tk.FLAT,
                 padx=10, pady=4).pack()
        
        # Hiển thị tin nhắn chào mừng
        welcome_msg = """👋 Xin chào! Tôi là trợ lý tài chính AI được hỗ trợ bởi Google Gemini.

🎯 Tôi có thể giúp bạn:
• Phân tích chi tiêu và đưa ra lời khuyên cụ thể
• Giải đáp thắc mắc về quản lý tài chính
• Đề xuất cách tiết kiệm hiệu quả
• Đánh giá xu hướng chi tiêu của bạn
• Hỗ trợ lập kế hoạch tài chính

💬 Hãy hỏi tôi bất cứ điều gì hoặc chọn câu hỏi gợi ý bên dưới! 😊"""
        
        self.display_message("Gemini AI", welcome_msg, "bot")
    
    def send_message(self):
        """Gửi tin nhắn đến ChatBot"""
        message = self.message_var.get().strip()
        if not message:
            return
        
        # Hiển thị tin nhắn người dùng
        self.display_message("Bạn", message, "user")
        self.message_var.set("")
        
        # Hiển thị "đang suy nghĩ..."
        self.chat_display.insert(tk.END, "\n🤔 Đang phân tích...\n", "system")
        self.chat_display.see(tk.END)
        self.chat_display.update()
        
        # Gọi ChatBot
        try:
            response = self.chatbot.ask_question(message)
        except Exception as e:
            response = f"❌ Lỗi: {str(e)}"
        
        # Xóa "đang suy nghĩ..."
        self.chat_display.delete("end-2l", "end-1l")
        
        # Hiển thị phản hồi
        self.display_message("Gemini AI", response, "bot")
    
    def send_suggestion(self, suggestion):
        """Gửi câu hỏi gợi ý"""
        self.message_var.set(suggestion)
        self.send_message()
    
    def reset_chat(self):
        """Reset cuộc trò chuyện"""
        if messagebox.askyesno("Xác nhận", 
            "Bạn có muốn bắt đầu cuộc trò chuyện mới?\n"
            "Lịch sử chat hiện tại sẽ bị xóa."):
            
            self.chatbot.clear_history()
            self.chat_display.delete(1.0, tk.END)
            
            welcome_msg = "🔄 Đã bắt đầu cuộc trò chuyện mới!\n\n" \
                         "💬 Hãy hỏi tôi bất cứ điều gì về tài chính cá nhân! 😊"
            self.display_message("Hệ thống", welcome_msg, "system")
    
    def open_ai_auto_input(self):
        """Mở cửa sổ nhập liệu tự động bằng AI"""
        if not AI_AUTO_INPUT_AVAILABLE or not self.ai_auto_input or not self.ai_auto_input.is_available():
            messagebox.showinfo(
                "Chức năng chưa sẵn sàng",
                "Tính năng Nhập bằng AI chưa được cấu hình.\n\n"
                "Vui lòng kiểm tra:\n"
                "1. Đã cài đặt google-generativeai\n"
                "2. Đã cấu hình GOOGLE_API_KEY_AUTO_INPUT trong config.py"
            )
            return
        
        # Tạo cửa sổ
        ai_window = tk.Toplevel(self.root)
        ai_window.title("🤖 Nhập Giao Dịch Bằng AI")
        ai_window.geometry("700x600")
        ai_window.configure(bg="#f5f5f5")
        
        # Frame chính
        main_frame = tk.Frame(ai_window, bg="white", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Tiêu đề
        title_frame = tk.Frame(main_frame, bg="#FF5722", pady=12)
        title_frame.pack(fill=tk.X)
        
        tk.Label(title_frame, text="🤖 Nhập Giao Dịch Bằng AI",
                bg="#FF5722", fg="white",
                font=("Arial", 14, "bold")).pack()
        
        tk.Label(title_frame, text="Chỉ cần nói, AI sẽ tự động thêm giao dịch",
                bg="#FF5722", fg="white",
                font=("Arial", 9)).pack()
        
        # Hướng dẫn
        guide_frame = tk.LabelFrame(main_frame, text="💡 Hướng dẫn",
                                    bg="white", font=("Arial", 10, "bold"),
                                    fg="#FF5722")
        guide_frame.pack(fill=tk.X, pady=10)
        
        guide_text = """
📝 Cách sử dụng:
• Gõ văn bản mô tả giao dịch (VD: "Vừa mua cà phê 50k")
• AI sẽ tự động phân tích và đề xuất giao dịch
• Xác nhận để thêm vào hệ thống

✅ Ví dụ:
• "Nhận lương 15 triệu"
• "Mua cafe 45 nghìn"
• "Hôm qua ăn phở 50k"
• "Đổ xăng 200k"
• "Nộp tiền nhà 5 triệu"
"""
        
        tk.Label(guide_frame, text=guide_text, bg="white",
                font=("Arial", 9), justify=tk.LEFT).pack(padx=10, pady=5)
        
        # Khu vực chat
        chat_frame = tk.Frame(main_frame, bg="white")
        chat_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.ai_chat_display = tk.Text(chat_frame, wrap=tk.WORD,
                                       font=("Arial", 10),
                                       bg="#f9f9f9", relief=tk.FLAT,
                                       padx=10, pady=10, height=15)
        self.ai_chat_display.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = tk.Scrollbar(chat_frame, command=self.ai_chat_display.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.ai_chat_display.config(yscrollcommand=scrollbar.set)
        
        # Tag format
        self.ai_chat_display.tag_config("user", foreground="#1976D2", font=("Arial", 10, "bold"))
        self.ai_chat_display.tag_config("ai", foreground="#FF5722", font=("Arial", 10, "bold"))
        self.ai_chat_display.tag_config("success", foreground="#4CAF50", font=("Arial", 10, "bold"))
        self.ai_chat_display.tag_config("error", foreground="#F44336", font=("Arial", 10))
        
        # Frame nhập
        input_frame = tk.Frame(main_frame, bg="white")
        input_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.ai_message_var = tk.StringVar()
        message_entry = tk.Entry(input_frame, textvariable=self.ai_message_var,
                                font=("Arial", 11), relief=tk.FLAT,
                                bg="#f0f0f0", bd=2)
        message_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10), ipady=5)
        message_entry.bind('<Return>', lambda e: self.process_ai_input())
        message_entry.focus()
        
        tk.Button(input_frame, text="Phân tích ➤",
                 command=self.process_ai_input,
                 bg="#FF5722", fg="white",
                 font=("Arial", 10, "bold"),
                 cursor="hand2", padx=20, relief=tk.FLAT).pack(side=tk.LEFT)
        
        # Biến lưu giao dịch đang chờ
        self.pending_transaction = None
        
        # Tin nhắn chào mừng
        welcome = "👋 Xin chào! Hãy nói cho tôi biết giao dịch của bạn.\n\n" \
                 "VD: 'Vừa mua cà phê 50k' hoặc 'Nhận lương 15 triệu'"
        
        self.ai_chat_display.insert(tk.END, "🤖 AI Assistant:\n", "ai")
        self.ai_chat_display.insert(tk.END, welcome + "\n")
    
    def process_ai_input(self):
        """Xử lý input từ người dùng với AI"""
        message = self.ai_message_var.get().strip()
        if not message:
            return
        
        # Hiển thị tin nhắn người dùng
        self.ai_chat_display.insert(tk.END, "\n" + "─" * 70 + "\n")
        self.ai_chat_display.insert(tk.END, "👤 Bạn:\n", "user")
        self.ai_chat_display.insert(tk.END, message + "\n")
        self.ai_message_var.set("")
        
        # Hiển thị đang xử lý
        self.ai_chat_display.insert(tk.END, "\n🤖 Đang phân tích...\n", "ai")
        self.ai_chat_display.see(tk.END)
        self.ai_chat_display.update()
        
        # Lấy danh mục có sẵn
        available_categories = self.get_available_categories()
        
        # Gọi AI phân tích
        result = self.ai_auto_input.parse_transaction(message, available_categories)
        
        # Xóa "đang phân tích"
        self.ai_chat_display.delete("end-2l", "end-1l")
        
        if not result:
            self.ai_chat_display.insert(tk.END, "\n❌ Lỗi: Không thể phân tích tin nhắn.\n", "error")
            return
        
        if not result.get('is_transaction', False):
            self.ai_chat_display.insert(tk.END, "\n🤖 AI:\n", "ai")
            self.ai_chat_display.insert(tk.END,
                "Xin lỗi, tôi không nhận ra đây là giao dịch tài chính.\n"
                "Vui lòng mô tả rõ hơn (VD: 'Mua cafe 50k' hoặc 'Nhận lương 10 triệu')\n")
            return
        
        # Hiển thị thông tin giao dịch
        self.ai_chat_display.insert(tk.END, "\n🤖 AI:\n", "ai")
        confirm_msg = self.ai_auto_input.confirm_transaction(result)
        self.ai_chat_display.insert(tk.END, confirm_msg + "\n")
        
        # Lưu giao dịch đang chờ
        self.pending_transaction = result
        
        # Tạo frame nút xác nhận
        button_frame = tk.Frame(self.ai_chat_display, bg="#f9f9f9")
        self.ai_chat_display.window_create(tk.END, window=button_frame)
        
        tk.Button(button_frame, text="✅ Thêm giao dịch",
                 command=lambda: self.confirm_add_transaction(result),
                 bg="#4CAF50", fg="white",
                 font=("Arial", 9, "bold"),
                 cursor="hand2", padx=15, pady=5).pack(side=tk.LEFT, padx=5)
        
        tk.Button(button_frame, text="❌ Hủy",
                 command=self.cancel_transaction,
                 bg="#F44336", fg="white",
                 font=("Arial", 9, "bold"),
                 cursor="hand2", padx=15, pady=5).pack(side=tk.LEFT, padx=5)
        
        self.ai_chat_display.insert(tk.END, "\n")
        self.ai_chat_display.see(tk.END)
    
    def get_available_categories(self):
        """Lấy danh sách danh mục có sẵn"""
        categories = {'income': [], 'expense': []}
        
        self.cursor.execute('SELECT name FROM categories WHERE type = "income"')
        categories['income'] = [row[0] for row in self.cursor.fetchall()]
        
        self.cursor.execute('SELECT name FROM categories WHERE type = "expense"')
        categories['expense'] = [row[0] for row in self.cursor.fetchall()]
        
        return categories
    
    def confirm_add_transaction(self, transaction):
        """Xác nhận và thêm giao dịch"""
        try:
            # Thêm vào database
            self.cursor.execute('''
                INSERT INTO transactions (type, category, amount, description, date, user_id)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                transaction['type'],
                transaction['category'],
                transaction['amount'],
                transaction['description'],
                transaction['date'],
                self.user_id
            ))
            
            self.conn.commit()
            
            # Hiển thị thành công
            self.ai_chat_display.insert(tk.END, "\n✅ Thành công!\n", "success")
            self.ai_chat_display.insert(tk.END,
                f"Đã thêm giao dịch {transaction['amount']:,.0f} VNĐ vào hệ thống.\n\n"
                "Bạn có thể tiếp tục nhập giao dịch khác! 😊\n")
            
            # Reset
            self.pending_transaction = None
            
            # Cập nhật danh sách giao dịch
            self.load_transactions()
            
        except Exception as e:
            self.ai_chat_display.insert(tk.END, f"\n❌ Lỗi: {str(e)}\n", "error")
    
    def cancel_transaction(self):
        """Hủy giao dịch"""
        self.ai_chat_display.insert(tk.END, "\n❌ Đã hủy giao dịch.\n")
        self.ai_chat_display.insert(tk.END, "Bạn có thể thử lại với mô tả khác.\n")
        self.pending_transaction = None
    
    def open_receipt_ocr(self):
        """Mở cửa sổ quét hóa đơn"""
        if not self.receipt_ocr:
            messagebox.showwarning("Chưa cấu hình",
                                 "Tính năng Quét Hóa Đơn chưa được cấu hình.\n\n"
                                 "Vui lòng:\n"
                                 "1. Cài đặt: pip install pillow google-generativeai\n"
                                 "2. Cấu hình API Key trong file config.py")
            return
        
        ocr_window = tk.Toplevel(self.root)
        ocr_window.title("📷 Quét Hóa Đơn - AI OCR")
        ocr_window.geometry("700x650")
        ocr_window.configure(bg="#f5f5f5")
        
        # Frame chính
        main_frame = tk.Frame(ocr_window, bg="white", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Tiêu đề
        title_frame = tk.Frame(main_frame, bg="#FF9800", pady=12)
        title_frame.pack(fill=tk.X)
        
        tk.Label(title_frame, text="📷 Quét Hóa Đơn Tự Động",
                bg="#FF9800", fg="white",
                font=("Arial", 14, "bold")).pack()
        
        tk.Label(title_frame, text="AI sẽ tự động đọc và tạo giao dịch từ ảnh hóa đơn",
                bg="#FF9800", fg="white",
                font=("Arial", 9)).pack()
        
        # Frame hiển thị ảnh
        image_frame = tk.LabelFrame(main_frame, text="📸 Hình ảnh hóa đơn",
                                   bg="white", font=("Arial", 10, "bold"),
                                   fg="#FF9800", padx=10, pady=10)
        image_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Label hiển thị ảnh
        self.ocr_image_label = tk.Label(image_frame, 
                                       text="📁 Chưa chọn ảnh\n\nClick 'Chọn Ảnh' để tải hóa đơn",
                                       bg="#f9f9f9", fg="#666",
                                       font=("Arial", 11),
                                       width=60, height=12,
                                       relief=tk.FLAT, bd=1)
        self.ocr_image_label.pack(fill=tk.BOTH, expand=True)
        
        # Frame thông tin
        info_frame = tk.LabelFrame(main_frame, text="📋 Thông tin trích xuất",
                                  bg="white", font=("Arial", 10, "bold"),
                                  fg="#FF9800")
        info_frame.pack(fill=tk.X, pady=10)
        
        self.ocr_info_text = tk.Text(info_frame, height=8, wrap=tk.WORD,
                                     font=("Arial", 10), bg="#f9f9f9",
                                     relief=tk.FLAT, padx=10, pady=10)
        self.ocr_info_text.pack(fill=tk.X, padx=5, pady=5)
        self.ocr_info_text.insert(1.0, "Chưa có thông tin. Vui lòng tải ảnh hóa đơn lên.")
        self.ocr_info_text.config(state=tk.DISABLED)
        
        # Frame buttons
        button_frame = tk.Frame(main_frame, bg="white")
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Biến lưu đường dẫn ảnh và dữ liệu
        self.current_image_path = None
        self.current_receipt_data = None
        
        # Nút chọn ảnh
        select_btn = tk.Button(button_frame, text="📁 Chọn Ảnh",
                              command=lambda: self.select_receipt_image(ocr_window),
                              bg="#2196F3", fg="white",
                              font=("Arial", 10, "bold"),
                              cursor="hand2", padx=20, pady=8)
        select_btn.pack(side=tk.LEFT, padx=5)
        
        # Nút quét
        self.scan_btn = tk.Button(button_frame, text="🔍 Quét Hóa Đơn",
                                 command=self.scan_receipt,
                                 bg="#FF9800", fg="white",
                                 font=("Arial", 10, "bold"),
                                 cursor="hand2", padx=20, pady=8,
                                 state=tk.DISABLED)
        self.scan_btn.pack(side=tk.LEFT, padx=5)
        
        # Nút thêm giao dịch
        self.add_receipt_btn = tk.Button(button_frame, text="✅ Thêm Giao Dịch",
                                        command=self.add_receipt_transaction,
                                        bg="#4CAF50", fg="white",
                                        font=("Arial", 10, "bold"),
                                        cursor="hand2", padx=20, pady=8,
                                        state=tk.DISABLED)
        self.add_receipt_btn.pack(side=tk.LEFT, padx=5)
        
        # Hướng dẫn
        help_frame = tk.Frame(main_frame, bg="#E3F2FD", relief=tk.FLAT, bd=1)
        help_frame.pack(fill=tk.X, pady=(10, 0))
        
        help_text = """💡 Hướng dẫn:
1. Click 'Chọn Ảnh' để chọn ảnh hóa đơn (JPG, PNG)
2. Click 'Quét Hóa Đơn' để AI phân tích
3. Kiểm tra thông tin và click 'Thêm Giao Dịch'

📝 Lưu ý: Ảnh nên rõ nét, đủ sáng để AI đọc tốt nhất"""
        
        tk.Label(help_frame, text=help_text, bg="#E3F2FD", fg="#1976D2",
                font=("Arial", 9), justify=tk.LEFT, padx=10, pady=8).pack()
    
    def select_receipt_image(self, window):
        """Chọn ảnh hóa đơn"""
        file_path = filedialog.askopenfilename(
            parent=window,
            title="Chọn ảnh hóa đơn",
            filetypes=[
                ("Image files", "*.jpg *.jpeg *.png *.bmp *.gif"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.current_image_path = file_path
            
            # Hiển thị preview ảnh
            try:
                from PIL import Image, ImageTk
                
                img = Image.open(file_path)
                
                # Resize để hiển thị
                img.thumbnail((500, 280), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                
                self.ocr_image_label.config(image=photo, text="")
                self.ocr_image_label.image = photo  # Keep reference
                
                # Enable nút quét
                self.scan_btn.config(state=tk.NORMAL)
                
                # Update info
                self.ocr_info_text.config(state=tk.NORMAL)
                self.ocr_info_text.delete(1.0, tk.END)
                filename = file_path.split('/')[-1] if '/' in file_path else file_path.split('\\')[-1]
                self.ocr_info_text.insert(1.0, f"✅ Đã chọn ảnh: {filename}\n\nClick 'Quét Hóa Đơn' để AI phân tích.")
                self.ocr_info_text.config(state=tk.DISABLED)
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể tải ảnh: {str(e)}")
    
    def scan_receipt(self):
        """Quét hóa đơn bằng AI"""
        if not self.current_image_path:
            messagebox.showwarning("Chưa chọn ảnh", "Vui lòng chọn ảnh hóa đơn trước!")
            return
        
        # Hiển thị loading
        self.ocr_info_text.config(state=tk.NORMAL)
        self.ocr_info_text.delete(1.0, tk.END)
        self.ocr_info_text.insert(1.0, "🔍 Đang quét hóa đơn...\nVui lòng đợi...")
        self.ocr_info_text.config(state=tk.DISABLED)
        self.ocr_info_text.update()
        
        # Disable buttons
        self.scan_btn.config(state=tk.DISABLED)
        
        try:
            # Gọi API OCR
            result = self.receipt_ocr.extract_receipt_info(self.current_image_path)
            
            if result['success']:
                data = result['data']
                self.current_receipt_data = data
                
                # Hiển thị kết quả
                info_text = f"""✅ Quét thành công!

📋 Thông tin trích xuất:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
💰 Số tiền: {data['amount']:,.0f} VNĐ
📁 Danh mục: {data['category']}
📝 Mô tả: {data['description']}
📅 Ngày: {data['date']}
🏪 Cửa hàng: {data.get('merchant', 'N/A')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

✅ Kiểm tra thông tin và click 'Thêm Giao Dịch' để lưu."""
                
                self.ocr_info_text.config(state=tk.NORMAL)
                self.ocr_info_text.delete(1.0, tk.END)
                self.ocr_info_text.insert(1.0, info_text)
                self.ocr_info_text.config(state=tk.DISABLED)
                
                # Enable nút thêm
                self.add_receipt_btn.config(state=tk.NORMAL)
                
            else:
                error_msg = f"❌ Lỗi: {result['error']}\n\nVui lòng thử lại với ảnh khác hoặc ảnh rõ hơn."
                self.ocr_info_text.config(state=tk.NORMAL)
                self.ocr_info_text.delete(1.0, tk.END)
                self.ocr_info_text.insert(1.0, error_msg)
                self.ocr_info_text.config(state=tk.DISABLED)
                
                messagebox.showerror("Lỗi quét hóa đơn", result['error'])
        
        except Exception as e:
            error_msg = f"❌ Có lỗi xảy ra: {str(e)}"
            self.ocr_info_text.config(state=tk.NORMAL)
            self.ocr_info_text.delete(1.0, tk.END)
            self.ocr_info_text.insert(1.0, error_msg)
            self.ocr_info_text.config(state=tk.DISABLED)
            
            messagebox.showerror("Lỗi", str(e))
        
        finally:
            # Enable lại nút quét
            self.scan_btn.config(state=tk.NORMAL)
    
    def add_receipt_transaction(self):
        """Thêm giao dịch từ hóa đơn đã quét"""
        if not self.current_receipt_data:
            messagebox.showwarning("Chưa có dữ liệu", "Vui lòng quét hóa đơn trước!")
            return
        
        try:
            data = self.current_receipt_data
            
            # Thêm vào database
            self.cursor.execute('''
                INSERT INTO transactions (user_id, type, amount, category, description, date)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (self.user_id, data['type'], data['amount'], 
                  data['category'], data['description'], data['date']))
            
            self.conn.commit()
            
            # Cập nhật danh sách
            self.load_transactions()
            self.check_budget_warning()
            
            # Hiển thị thông báo
            messagebox.showinfo("Thành công", 
                              f"✅ Đã thêm giao dịch:\n\n"
                              f"💰 {data['amount']:,.0f} VNĐ\n"
                              f"📁 {data['category']}\n"
                              f"📝 {data['description']}")
            
            # Reset
            self.current_receipt_data = None
            self.current_image_path = None
            self.add_receipt_btn.config(state=tk.DISABLED)
            
            # Clear info
            self.ocr_info_text.config(state=tk.NORMAL)
            self.ocr_info_text.delete(1.0, tk.END)
            self.ocr_info_text.insert(1.0, "✅ Đã thêm giao dịch thành công!\n\nChọn ảnh khác để tiếp tục.")
            self.ocr_info_text.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể thêm giao dịch: {str(e)}")
    
    def update_gold_price(self):
        """Cập nhật giá vàng hiện tại"""
        if not GOLD_PRICE_AVAILABLE or not self.gold_api:
            self.gold_price_label.config(
                text="⚠️ Chưa khả dụng\n\nCài: pip install\nrequests",
                fg="#999"
            )
            return
        
        try:
            # Hiển thị đang tải
            self.gold_price_label.config(text="⏳ Đang tải...", fg="#666")
            self.gold_price_label.update()
            
            # Gọi API
            result = self.gold_api.get_current_price()
            
            if result['success']:
                price_per_gram = result['price_per_gram']
                change_24h = result['change_24h']
                timestamp = result['timestamp']
                is_reference = 'note' in result
                
                # Lấy giá USD/ounce
                price_usd_ounce = result['price'] / 24500  # Chuyển ngược lại từ VND sang USD
                
                # Icon và màu cho thay đổi
                if change_24h > 0:
                    change_icon = "📈"
                    change_color = "#4CAF50"
                elif change_24h < 0:
                    change_icon = "📉"
                    change_color = "#F44336"
                else:
                    change_icon = "➡️"
                    change_color = "#FF9800"
                
                # Format text ngắn gọn - hiển thị USD
                price_text = f"""💎 XAU

${price_usd_ounce:,.0f}"""
                
                if not is_reference and change_24h != 0:
                    price_text += f"\n{change_icon}{abs(change_24h):.1f}%"
                
                price_text += f"\n\n{timestamp.strftime('%H:%M')}"
                
                self.gold_price_label.config(text=price_text, fg=change_color)
            else:
                # Lỗi ngắn gọn
                self.gold_price_label.config(
                    text=f"❌ Lỗi\n\nThử lại",
                    fg="#F44336"
                )
        
        except Exception as e:
            self.gold_price_label.config(
                text=f"❌ Lỗi",
                fg="#F44336"
            )
    
    def update_btc_price(self):
        """Cập nhật giá Bitcoin hiện tại"""
        try:
            # Hiển thị đang tải
            self.btc_price_label.config(text="⏳", fg="#666")
            self.btc_price_label.update()
            
            # Gọi API CoinGecko (miễn phí, không cần key)
            import requests
            url = "https://api.coingecko.com/api/v3/simple/price?ids=bitcoin&vs_currencies=usd&include_24hr_change=true"
            
            response = requests.get(url, timeout=5)
            
            if response.status_code == 200:
                data = response.json()
                price_usd = data['bitcoin']['usd']
                change_24h = data['bitcoin'].get('usd_24h_change', 0)
                
                # Icon và màu cho thay đổi
                if change_24h > 0:
                    change_icon = "📈"
                    change_color = "#4CAF50"
                elif change_24h < 0:
                    change_icon = "📉"
                    change_color = "#F44336"
                else:
                    change_icon = "➡️"
                    change_color = "#F7931A"
                
                # Format text ngắn gọn
                from datetime import datetime
                price_text = f"""₿ BTC

${price_usd:,.0f}"""
                
                if change_24h != 0:
                    price_text += f"\n{change_icon}{abs(change_24h):.1f}%"
                
                price_text += f"\n\n{datetime.now().strftime('%H:%M')}"
                
                self.btc_price_label.config(text=price_text, fg=change_color)
            else:
                self.btc_price_label.config(
                    text=f"❌ Lỗi\n\nThử lại",
                    fg="#F44336"
                )
        
        except Exception as e:
            self.btc_price_label.config(
                text=f"❌ Lỗi",
                fg="#F44336"
            )
    
    def schedule_btc_price_update(self):
        """Lên lịch cập nhật giá Bitcoin tự động mỗi 5 phút"""
        # Auto-refresh sau 5 phút (300000 ms)
        self.root.after(300000, self.update_btc_price)
        self.root.after(300000, self.schedule_btc_price_update)
    
    def schedule_gold_price_update(self):
        """Lên lịch cập nhật giá vàng tự động mỗi 5 phút"""
        # Auto-refresh sau 5 phút (300000 ms)
        self.root.after(300000, self.update_gold_price)
        self.root.after(300000, self.schedule_gold_price_update)

    def display_message(self, sender, message, tag):
        """Hiển thị tin nhắn trong chat"""
        current_time = datetime.now().strftime("%H:%M")
        
        self.chat_display.insert(tk.END, f"\n{'─' * 80}\n")
        self.chat_display.insert(tk.END, f"{sender}", tag)
        self.chat_display.insert(tk.END, f" • {current_time}\n", "time")
        self.chat_display.insert(tk.END, f"{message}\n")
        self.chat_display.see(tk.END)

def hash_password(password):
    """Mã hóa mật khẩu bằng SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()

class LoginWindow:
    """Màn hình đăng nhập"""
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Đăng Nhập - Quản Lý Chi Tiêu")
        self.root.geometry("450x600")
        self.root.configure(bg="#f0f0f0")
        self.root.resizable(False, False)
        
        # Kết nối database
        self.conn = sqlite3.connect('finance.db')
        self.cursor = self.conn.cursor()
        
        # Khởi tạo bảng users nếu chưa có
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password TEXT NOT NULL
            )
        ''')
        self.conn.commit()
        
        self.user_id = None
        self.create_login_widgets()
        
        # Căn giữa cửa sổ
        self.center_window()
        
    def center_window(self):
        """Căn giữa cửa sổ trên màn hình"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_login_widgets(self):
        """Tạo giao diện đăng nhập"""
        # Frame chính
        main_frame = tk.Frame(self.root, bg="white", padx=40, pady=40)
        main_frame.pack(expand=True, fill=tk.BOTH, padx=20, pady=20)
        
        # Tiêu đề
        title_label = tk.Label(main_frame, text="ĐĂNG NHẬP",
                              bg="white", fg="#1a237e",
                              font=("Arial", 24, "bold"))
        title_label.pack(pady=(0, 10))
        
        subtitle_label = tk.Label(main_frame, text="Quản Lý Chi Tiêu Cá Nhân",
                                 bg="white", fg="#666",
                                 font=("Arial", 11))
        subtitle_label.pack(pady=(0, 30))
        
        # Username
        tk.Label(main_frame, text="Tên đăng nhập:",
                bg="white", font=("Arial", 10)).pack(anchor="w", pady=(10, 5))
        self.username_var = tk.StringVar()
        username_entry = tk.Entry(main_frame, textvariable=self.username_var,
                                  font=("Arial", 11), width=30)
        username_entry.pack(pady=(0, 15))
        username_entry.focus()
        
        # Password
        tk.Label(main_frame, text="Mật khẩu:",
                bg="white", font=("Arial", 10)).pack(anchor="w", pady=(10, 5))
        self.password_var = tk.StringVar()
        password_entry = tk.Entry(main_frame, textvariable=self.password_var,
                                 font=("Arial", 11), show="•", width=30)
        password_entry.pack(pady=(0, 25))
        password_entry.bind('<Return>', lambda e: self.login())
        
        # Nút đăng nhập
        login_btn = tk.Button(main_frame, text="Đăng Nhập",
                             command=self.login,
                             bg="#4CAF50", fg="white",
                             font=("Arial", 12, "bold"),
                             cursor="hand2", width=25, pady=10)
        login_btn.pack(pady=(0, 15))
        
        # Đường phân cách
        separator = tk.Frame(main_frame, height=1, bg="#ddd")
        separator.pack(fill=tk.X, pady=20)
        
        # Nút đăng ký
        tk.Label(main_frame, text="Chưa có tài khoản?",
                bg="white", fg="#666",
                font=("Arial", 10)).pack()
        
        register_btn = tk.Button(main_frame, text="Đăng Ký Ngay",
                                command=self.open_register,
                                bg="#2196F3", fg="white",
                                font=("Arial", 11, "bold"),
                                cursor="hand2", width=25, pady=8)
        register_btn.pack(pady=(10, 0))
    
    def login(self):
        """Xử lý đăng nhập"""
        username = self.username_var.get().strip()
        password = self.password_var.get()
        
        if not username or not password:
            messagebox.showerror("Lỗi", "Vui lòng nhập đầy đủ thông tin!")
            return
        
        # Mã hóa mật khẩu
        hashed_password = hash_password(password)
        
        # Kiểm tra trong database
        self.cursor.execute('SELECT id FROM users WHERE username = ? AND password = ?',
                          (username, hashed_password))
        result = self.cursor.fetchone()
        
        if result:
            self.user_id = result[0]
            messagebox.showinfo("Thành công", f"Chào mừng {username}!")
            self.conn.close()
            self.root.destroy()
            self.open_main_app()
        else:
            messagebox.showerror("Lỗi", "Tên đăng nhập hoặc mật khẩu không đúng!")
    
    def open_register(self):
        """Mở màn hình đăng ký"""
        RegisterWindow(self.root)
    
    def open_main_app(self):
        """Mở ứng dụng chính"""
        root = tk.Tk()
        app = FinanceManager(root, self.user_id)
        root.mainloop()
    
    def run(self):
        """Chạy ứng dụng"""
        self.root.mainloop()
        return self.user_id

class RegisterWindow:
    """Màn hình đăng ký"""
    def __init__(self, parent):
        self.window = tk.Toplevel(parent)
        self.window.title("Đăng Ký - Quản Lý Chi Tiêu")
        self.window.geometry("400x550")
        self.window.configure(bg="#f0f0f0")
        self.window.resizable(False, False)
        self.window.grab_set()  # Modal window
        
        # Kết nối database
        self.conn = sqlite3.connect('finance.db')
        self.cursor = self.conn.cursor()
        
        self.create_register_widgets()
        self.center_window()
    
    def center_window(self):
        """Căn giữa cửa sổ"""
        self.window.update_idletasks()
        width = self.window.winfo_width()
        height = self.window.winfo_height()
        x = (self.window.winfo_screenwidth() // 2) - (width // 2)
        y = (self.window.winfo_screenheight() // 2) - (height // 2)
        self.window.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_register_widgets(self):
        """Tạo giao diện đăng ký"""
        # Frame chính
        main_frame = tk.Frame(self.window, bg="white", padx=40, pady=40)
        main_frame.pack(expand=True, fill=tk.BOTH, padx=20, pady=20)
        
        # Tiêu đề
        title_label = tk.Label(main_frame, text="ĐĂNG KÝ TÀI KHOẢN",
                              bg="white", fg="#1a237e",
                              font=("Arial", 22, "bold"))
        title_label.pack(pady=(0, 10))
        
        subtitle_label = tk.Label(main_frame, text="Tạo tài khoản mới để bắt đầu",
                                 bg="white", fg="#666",
                                 font=("Arial", 10))
        subtitle_label.pack(pady=(0, 30))
        
        # Username
        tk.Label(main_frame, text="Tên đăng nhập:",
                bg="white", font=("Arial", 10)).pack(anchor="w", pady=(10, 5))
        self.username_var = tk.StringVar()
        username_entry = tk.Entry(main_frame, textvariable=self.username_var,
                                  font=("Arial", 11), width=30)
        username_entry.pack(pady=(0, 15))
        username_entry.focus()
        
        # Password
        tk.Label(main_frame, text="Mật khẩu:",
                bg="white", font=("Arial", 10)).pack(anchor="w", pady=(10, 5))
        self.password_var = tk.StringVar()
        password_entry = tk.Entry(main_frame, textvariable=self.password_var,
                                 font=("Arial", 11), show="•", width=30)
        password_entry.pack(pady=(0, 15))
        
        # Confirm Password
        tk.Label(main_frame, text="Xác nhận mật khẩu:",
                bg="white", font=("Arial", 10)).pack(anchor="w", pady=(10, 5))
        self.confirm_password_var = tk.StringVar()
        confirm_entry = tk.Entry(main_frame, textvariable=self.confirm_password_var,
                                font=("Arial", 11), show="•", width=30)
        confirm_entry.pack(pady=(0, 25))
        confirm_entry.bind('<Return>', lambda e: self.register())
        
        # Nút đăng ký
        register_btn = tk.Button(main_frame, text="Đăng Ký",
                                command=self.register,
                                bg="#4CAF50", fg="white",
                                font=("Arial", 12, "bold"),
                                cursor="hand2", width=25, pady=10)
        register_btn.pack(pady=(0, 15))
        
        # Nút hủy
        cancel_btn = tk.Button(main_frame, text="Hủy",
                              command=self.window.destroy,
                              bg="#9E9E9E", fg="white",
                              font=("Arial", 11),
                              cursor="hand2", width=25, pady=8)
        cancel_btn.pack()
    
    def register(self):
        """Xử lý đăng ký"""
        username = self.username_var.get().strip()
        password = self.password_var.get()
        confirm_password = self.confirm_password_var.get()
        
        # Kiểm tra dữ liệu
        if not username or not password or not confirm_password:
            messagebox.showerror("Lỗi", "Vui lòng nhập đầy đủ thông tin!")
            return
        
        if len(username) < 3:
            messagebox.showerror("Lỗi", "Tên đăng nhập phải có ít nhất 3 ký tự!")
            return
        
        if len(password) < 6:
            messagebox.showerror("Lỗi", "Mật khẩu phải có ít nhất 6 ký tự!")
            return
        
        if password != confirm_password:
            messagebox.showerror("Lỗi", "Mật khẩu xác nhận không khớp!")
            return
        
        # Kiểm tra username đã tồn tại chưa
        self.cursor.execute('SELECT COUNT(*) FROM users WHERE username = ?', (username,))
        if self.cursor.fetchone()[0] > 0:
            messagebox.showerror("Lỗi", "Tên đăng nhập đã tồn tại!")
            return
        
        try:
            # Mã hóa mật khẩu
            hashed_password = hash_password(password)
            
            # Thêm vào database
            self.cursor.execute('INSERT INTO users (username, password) VALUES (?, ?)',
                              (username, hashed_password))
            self.conn.commit()
            
            messagebox.showinfo("Thành công", 
                              f"Đăng ký thành công!\nTài khoản: {username}\n"
                              "Vui lòng đăng nhập để tiếp tục.")
            
            self.conn.close()
            self.window.destroy()
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra: {str(e)}")

def main():
    # Hiển thị màn hình đăng nhập
    login_window = LoginWindow()
    user_id = login_window.run()

if __name__ == "__main__":
    main()

